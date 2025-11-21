# views.py - Updated with admin user management

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login ,logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import CustomUser, UserActivityLog
from .forms import UserCreationForm, UserUpdateForm
from django.db.models import Q
from courses.models import BatchEnrollment

def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages
from courses.models import  DeviceSession
import hashlib

# userss/views.py

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from courses.models import StudentDeviceLimit, DeviceSession, StudentLoginLog
import json

def get_client_ip(request):
    """Get real IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip




def user_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        device_fingerprint_id = request.POST.get("device_fingerprint")  # From JS
        device_name = request.POST.get("device_name", "Unknown Device")
        device_info_raw = request.POST.get("device_info", "{}")
        
        print(f"🔐 Login attempt: {username}")
        print(f"📱 Device ID: {device_fingerprint_id[:20] if device_fingerprint_id else 'MISSING'}...")
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            
            # ========== DEVICE FINGERPRINT CHECK FOR STUDENTS ==========
            if user.role == "student":
                
                if not device_fingerprint_id:
                    return render(request, "login.html", {
                        "error": "Device fingerprint missing. Please enable JavaScript and try again.",
                        "username": username
                    })
                
                # Get or create device limit
                device_limit, created = StudentDeviceLimit.objects.get_or_create(
                    student=user,
                    defaults={'max_devices': 2, 'is_active': True}
                )
                
                if not device_limit.is_active:
                    return render(request, "login.html", {
                        "error": "Your account is disabled. Contact admin.",
                        "username": username
                    })
                
                # Parse device info
                try:
                    device_info = json.loads(device_info_raw)
                except:
                    device_info = {}
                
                # Check if device already registered
                existing_device = DeviceSession.objects.filter(
                    device_id=device_fingerprint_id,
                    student_limit=device_limit
                ).first()
                
                if existing_device:
                    # Device exists - allow login
                    print(f"✅ Known device: {existing_device.device_name}")
                    existing_device.is_active = True
                    existing_device.last_login = timezone.now()
                    existing_device.save()
                    
                    current_device_session = existing_device
                
                else:
                    # New device - check limit
                    if device_limit.can_add_device():
                        # Create new device
                        current_device_session = DeviceSession.objects.create(
                            student_limit=device_limit,
                            device_id=device_fingerprint_id,
                            device_name=device_name,
                            device_info=device_info,
                            is_active=True
                        )
                        print(f"✅ New device registered: {device_name}")
                    
                    else:
                        # Device limit reached - BLOCK
                        active_devices = DeviceSession.objects.filter(
                            student_limit=device_limit,
                            is_active=True
                        ).order_by('-last_login')
                        
                        error_msg = (
                            f'❌ Device Limit Reached!\n\n'
                            f'Maximum devices allowed: {device_limit.max_devices}\n'
                            f'Active devices: {active_devices.count()}\n\n'
                            f'Please remove an existing device or contact admin.'
                        )
                        
                        print(f"🚫 LOGIN BLOCKED: Device limit exceeded")
                        
                        return render(request, "login.html", {
                            "error": error_msg,
                            "username": username,
                            "active_devices": active_devices,
                            "new_device_name": device_name,
                        })
                
                # Create login log
                login_log = StudentLoginLog.objects.create(
                    student=user,
                    device_session=current_device_session,
                    ip_address=get_client_ip(request),
                    device_info=device_name
                )
                
                # Store in session for logout tracking
                request.session['attendance_log_id'] = login_log.id
                request.session['device_fingerprint'] = device_fingerprint_id
            
            # ========== END DEVICE CHECK ==========
            
            # Login user
            login(request, user)
            
            print(f"✅ LOGIN SUCCESS: {username}")
            
            # Redirect based on role
            if user.role == "superadmin":
                return redirect("admin_dashboard")
            elif user.role == "instructor":
                return redirect("instructor_dashboard")
            elif user.role == "student":
                return redirect("student_dashboard")
            elif user.role == "webinar_user":
                return redirect("webinars:webinar_landing")
            else:
                return render(request, "login.html", {"error": "Role not match"})
        else:
            print(f"❌ LOGIN FAILED: Invalid credentials")
            return render(request, "login.html", {"error": "Invalid Credentials"})
    
    return render(request, 'login.html')




@login_required
def user_logout(request):
    """Logout with attendance tracking"""
    
    print(f"🚪 Logout: {request.user.username if request.user.is_authenticated else 'Anonymous'}")
    
    if request.user.is_authenticated and hasattr(request.user, 'role') and request.user.role == 'student':
        try:
            # Get log ID from session
            log_id = request.session.get('attendance_log_id')
            
            if log_id:
                log = StudentLoginLog.objects.get(id=log_id)
                if not log.logout_time:
                    log.logout_time = timezone.now()
                    log.calculate_duration()
                    print(f"✅ Session closed: Duration {log.session_duration} min")
            
            # Close any other active sessions
            active_sessions = StudentLoginLog.objects.filter(
                student=request.user,
                logout_time__isnull=True
            )
            
            for session in active_sessions:
                session.logout_time = timezone.now()
                session.calculate_duration()
                session.save()
                print(f"✅ Auto-closed session {session.id}")
                
        except Exception as e:
            print(f"❌ Logout error: {e}")
    
    logout(request)
    return redirect('user_login')

# Admin Dashboard
@login_required
def admin_dashboard(request):
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    # Get user statistics
    total_users = CustomUser.objects.count()
    total_students = CustomUser.objects.filter(role='student').count()
    total_instructors = CustomUser.objects.filter(role='instructor').count()
    total_admins = CustomUser.objects.filter(role='superadmin').count()
    
    # Add courses data for sidebar dropdown
    from courses.models import Course
    sidebar_courses = Course.objects.filter(is_active=True).order_by('course_code')[:15]
    
    context = {
        'total_users': total_users,
        'total_students': total_students,
        'total_instructors': total_instructors,
        'total_admins': total_admins,
        'sidebar_courses': sidebar_courses,
    }
    return render(request, 'admin_dashboard.html', context)

# User Management Views
@login_required
def manage_users(request):
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    course_filter = request.GET.get('course', '')
    batch_type_filter = request.GET.get('batch_type', '')
    
    users = CustomUser.objects.all()
    
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if role_filter:
        users = users.filter(role=role_filter)
    
    # Course filter for students
    if course_filter and role_filter == 'student':
        users = users.filter(enrollments__course_id=course_filter, enrollments__is_active=True).distinct()
    
    # Batch type filter for students
    if batch_type_filter and role_filter == 'student':
        users = users.filter(batch_enrollments__batch__batch_type=batch_type_filter, batch_enrollments__is_active=True).distinct()
    
    users = users.order_by('-date_joined')
    
    # Get student details with courses and batches
    students_data = []
    if role_filter == 'student' or not role_filter:
        for user in users:
            if user.role == 'student':
                # Get enrollments
                enrollments = Enrollment.objects.filter(
                    student=user, 
                    is_active=True
                ).select_related('course', 'course__category')
                
                # Get batch enrollments
                batch_enrollments = BatchEnrollment.objects.filter(
                    student=user,
                    is_active=True
                ).select_related('batch', 'batch__course', 'batch__instructor')
                
                students_data.append({
                    'user': user,
                    'enrollments': enrollments,
                    'batch_enrollments': batch_enrollments,
                })
    
    # Get all courses for filter dropdown
    all_courses = Course.objects.filter(is_active=True).order_by('title')
    
    context = {
        'users': users,
        'students_data': students_data,
        'search_query': search_query,
        'role_filter': role_filter,
        'course_filter': course_filter,
        'batch_type_filter': batch_type_filter,
        'role_choices': CustomUser.ROLE_CHOICES,
        'all_courses': all_courses,
        'batch_type_choices': [
            ('online', 'Online'),
            ('offline', 'Offline'),
            ('hybrid', 'Hybrid'),
        ],
    }
    return render(request, 'manage_users.html', context)

    
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import UserCreationForm

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import UserCreationForm

@login_required
def create_user(request):
    if request.user.role != 'superadmin':
        messages.error(request, 'You do not have permission to create users.')
        return redirect('user_login')
    
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'User {user.username} created successfully!')
            return redirect('manage_users')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UserCreationForm()
    
    return render(request, 'create_user.html', {'form': form})






@login_required
def edit_user(request, user_id):
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    user = get_object_or_404(CustomUser, id=user_id)
    
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, f'User {user.username} updated successfully!')
            return redirect('manage_users')
    else:
        form = UserUpdateForm(instance=user)
    
    return render(request, 'edit_user.html', {'form': form, 'user': user})

@login_required
def delete_user(request, user_id):
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    user = get_object_or_404(CustomUser, id=user_id)
    
    # Prevent self-deletion
    if user == request.user:
        messages.error(request, "You cannot delete your own account!")
        return redirect('manage_users')
    
    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f'User {username} deleted successfully!')
        return redirect('manage_users')
    
    return render(request, 'delete_user.html', {'user': user})

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from datetime import date, timedelta
from userss.models import CustomUser, EmailLog, EmailLimitSet
from courses.models import Course, Batch, Enrollment, BatchEnrollment
from exams.models import Exam, ExamAttempt
from zoom.models import BatchSession

@login_required
def user_details(request, user_id):
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    user = get_object_or_404(CustomUser, id=user_id)
    
    # Common data for all users
    context = {
        'user': user,
    }
    
    # Email Statistics (for all users)
    today = date.today()
    email_limit = EmailLimitSet.objects.filter(is_active=True).first()
    daily_limit = email_limit.email_limit_per_day if email_limit else 50
    
    # Emails sent TO this user
    today_emails = EmailLog.objects.filter(
        recipient_user=user,
        sent_date=today
    ).count()
    
    total_emails = EmailLog.objects.filter(
        recipient_user=user
    ).count()
    
    remaining_emails = max(0, daily_limit - today_emails)
    can_receive_email = remaining_emails > 0
    
    # Recent emails
    recent_emails = EmailLog.objects.filter(
        recipient_user=user
    ).select_related('sent_by', 'template_used').order_by('-sent_time')[:10]
    
    # Add email data to context
    context.update({
        'today_emails': today_emails,
        'total_emails': total_emails,
        'remaining_emails': remaining_emails,
        'can_receive_email': can_receive_email,
        'daily_limit': daily_limit,
        'recent_emails': recent_emails,
    })
    
    # Role-specific data
    if user.role == 'instructor':
        # Instructor Statistics
        instructor_data = get_instructor_statistics(user)
        context.update(instructor_data)
        
    elif user.role == 'student':
        # Student Statistics
        student_data = get_student_statistics(user)
        context.update(student_data)
    
    # Emails SENT by this user (if instructor or admin)
    if user.role in ['instructor', 'superadmin']:
        emails_sent_by_user = EmailLog.objects.filter(
            sent_by=user
        ).count()
        
        emails_sent_today = EmailLog.objects.filter(
            sent_by=user,
            sent_date=today
        ).count()
        
        context.update({
            'emails_sent_by_user': emails_sent_by_user,
            'emails_sent_today': emails_sent_today,
        })
    
    return render(request, 'user_details.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from datetime import date

@login_required
def instructor_user_details(request, user_id):
    # Only instructors can access this view
    if request.user.role != 'instructor':
        return redirect('user_login')
    
    # Instructor can only view their own details
    if request.user.id != user_id:
        return redirect('user_login')
    
    user = get_object_or_404(CustomUser, id=user_id)
    
    # Common data for the instructor
    context = {
        'user': user,
    }
    
    # Email Statistics
    today = date.today()
    email_limit = EmailLimitSet.objects.filter(is_active=True).first()
    daily_limit = email_limit.email_limit_per_day if email_limit else 50
    
    # Emails sent TO this instructor
    today_emails = EmailLog.objects.filter(
        recipient_user=user,
        sent_date=today
    ).count()
    
    total_emails = EmailLog.objects.filter(
        recipient_user=user
    ).count()
    
    remaining_emails = max(0, daily_limit - today_emails)
    can_receive_email = remaining_emails > 0
    
    # Recent emails
    recent_emails = EmailLog.objects.filter(
        recipient_user=user
    ).select_related('sent_by', 'template_used').order_by('-sent_time')[:10]
    
    # Add email data to context
    context.update({
        'today_emails': today_emails,
        'total_emails': total_emails,
        'remaining_emails': remaining_emails,
        'can_receive_email': can_receive_email,
        'daily_limit': daily_limit,
        'recent_emails': recent_emails,
    })
    
    # Instructor-specific statistics
    instructor_data = get_instructor_statistics(user)
    context.update(instructor_data)
    
    # Emails SENT by this instructor
    emails_sent_by_user = EmailLog.objects.filter(
        sent_by=user
    ).count()
    
    emails_sent_today = EmailLog.objects.filter(
        sent_by=user,
        sent_date=today
    ).count()
    
    context.update({
        'emails_sent_by_user': emails_sent_by_user,
        'emails_sent_today': emails_sent_today,
    })
    
    return render(request, 'user_details_instructor.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from datetime import date
from courses.models import *


@login_required
def instructor_view_student_profile(request, student_id):
    """Instructor student profile dekh sakta hai - bilkul user_details jaisa"""
    
    # Only instructor access kar sakta
    if request.user.role != 'instructor':
        return redirect('user_login')
    
    # Student get karo
    student = get_object_or_404(CustomUser, id=student_id, role='student')
    
    # Check: Is student enrolled in this instructor's course?
    if not Enrollment.objects.filter(student=student, course__instructor=request.user).exists():
        messages.error(request, "You don't have permission to view this student.")
        return redirect('instructor_student_management')
    
    # Common data
    context = {
        'user': student,
        'is_viewing_student': True,  # Template ke liye flag
    }
    
    # Email Statistics (bilkul user_details jaisa)
    today = date.today()
    email_limit = EmailLimitSet.objects.filter(is_active=True).first()
    daily_limit = email_limit.email_limit_per_day if email_limit else 50
    
    today_emails = EmailLog.objects.filter(
        recipient_user=student,
        sent_date=today
    ).count()
    
    total_emails = EmailLog.objects.filter(
        recipient_user=student
    ).count()
    
    remaining_emails = max(0, daily_limit - today_emails)
    can_receive_email = remaining_emails > 0
    
    recent_emails = EmailLog.objects.filter(
        recipient_user=student
    ).select_related('sent_by', 'template_used').order_by('-sent_time')[:10]
    
    context.update({
        'today_emails': today_emails,
        'total_emails': total_emails,
        'remaining_emails': remaining_emails,
        'can_receive_email': can_receive_email,
        'daily_limit': daily_limit,
        'recent_emails': recent_emails,
    })
    
    # Student Statistics (bas instructor ke courses ka data)
    student_data = get_student_statistics_for_instructor(student, request.user)
    context.update(student_data)
    
    return render(request, 'user_details_instructor.html', context)


def get_student_statistics_for_instructor(student, instructor):
    """Student ki statistics - sirf instructor ke courses ka"""
    
    # Enrollments
    enrollments = Enrollment.objects.filter(
        student=student,
        course__instructor=instructor
    ).select_related('course')
    
    # Batch Enrollments
    batch_enrollments = BatchEnrollment.objects.filter(
        student=student,
        batch__instructor=instructor
    ).select_related('batch', 'batch__course')
    
    # Exam Attempts
    exam_attempts = ExamAttempt.objects.filter(
        student=student,
        exam__created_by=instructor
    ).select_related('exam')
    
    completed_exams = exam_attempts.filter(status='completed', is_graded=True).count()
    passed_exams = exam_attempts.filter(status='completed', is_graded=True, is_passed=True).count()
    
    graded_attempts = exam_attempts.filter(is_graded=True)
    avg_percentage = round(graded_attempts.aggregate(avg=Avg('percentage'))['avg'] or 0, 2)
    
    return {
        # Courses
        'enrolled_courses': enrollments.order_by('-enrolled_at'),
        'total_enrollments': enrollments.count(),
        'active_enrollments': enrollments.filter(status='enrolled', is_active=True).count(),
        
        # Batches
        'enrolled_batches': batch_enrollments.order_by('-enrolled_at'),
        'total_batch_enrollments': batch_enrollments.count(),
        'active_batch_enrollments': batch_enrollments.filter(status='enrolled', is_active=True).count(),
        
        # Exams
        'recent_exam_attempts': exam_attempts.order_by('-started_at')[:5],
        'completed_exams': completed_exams,
        'passed_exams': passed_exams,
        'avg_percentage': avg_percentage,
    }

def get_instructor_statistics(user):
    """Get comprehensive statistics for instructor"""
    
    # Courses
    total_courses = Course.objects.filter(
        Q(instructor=user) | Q(co_instructors=user),
        is_active=True
    ).distinct().count()
    
    active_courses = Course.objects.filter(
        Q(instructor=user) | Q(co_instructors=user),
        status='published',
        is_active=True
    ).distinct().count()
    
    # Get all courses (main + co-instructor)
    all_courses = Course.objects.filter(
        Q(instructor=user) | Q(co_instructors=user),
        is_active=True
    ).distinct().select_related('category')[:5]  # Recent 5
    
    # Batches
    total_batches = Batch.objects.filter(
        instructor=user,
        is_active=True
    ).count()
    
    active_batches = Batch.objects.filter(
        instructor=user,
        status='active',
        is_active=True
    ).count()
    
    recent_batches = Batch.objects.filter(
        instructor=user,
        is_active=True
    ).select_related('course').order_by('-created_at')[:5]
    
    # Students (enrolled in instructor's courses)
    total_students = Enrollment.objects.filter(
        course__instructor=user,
        is_active=True
    ).values('student').distinct().count()
    
    # Get recent students
    recent_students = Enrollment.objects.filter(
        course__instructor=user,
        is_active=True
    ).select_related('student', 'course').order_by('-enrolled_at')[:10]
    
    # Sessions (from batches) - DETAILED
    total_sessions = BatchSession.objects.filter(
        batch__instructor=user
    ).count()
    
    upcoming_sessions = BatchSession.objects.filter(
        batch__instructor=user,
        status='scheduled',
        scheduled_date__gte=date.today()
    ).count()
    
    completed_sessions = BatchSession.objects.filter(
        batch__instructor=user,
        status='completed'
    ).count()
    
    live_sessions = BatchSession.objects.filter(
        batch__instructor=user,
        status='live'
    ).count()
    
    # ALL Sessions with details
    all_sessions = BatchSession.objects.filter(
        batch__instructor=user
    ).select_related('batch', 'batch__course').order_by('-scheduled_date', '-start_time')[:20]
    
    # Session stats by status
    session_stats = BatchSession.objects.filter(
        batch__instructor=user
    ).values('status').annotate(count=Count('id'))
    
    recent_sessions = BatchSession.objects.filter(
        batch__instructor=user
    ).select_related('batch').order_by('-scheduled_date', '-start_time')[:5]
    
    # Exams
    total_exams = Exam.objects.filter(
        created_by=user,
        is_active=True
    ).count()
    
    active_exams = Exam.objects.filter(
        created_by=user,
        status='published',
        is_active=True
    ).count()
    
    recent_exams = Exam.objects.filter(
        created_by=user,
        is_active=True
    ).order_by('-created_at')[:5]
    
    # Exam attempts on instructor's exams
    total_attempts = ExamAttempt.objects.filter(
        exam__created_by=user
    ).count()
    
    # Permissions (if applicable)
    instructor_permissions = []
    if hasattr(user, 'instructor_profile'):
        instructor_permissions = user.get_instructor_permissions()
    
    return {
        'total_courses': total_courses,
        'active_courses': active_courses,
        'all_courses': all_courses,
        
        'total_batches': total_batches,
        'active_batches': active_batches,
        'recent_batches': recent_batches,
        
        'total_students': total_students,
        'recent_students': recent_students,
        
        'total_sessions': total_sessions,
        'upcoming_sessions': upcoming_sessions,
        'completed_sessions': completed_sessions,
        'live_sessions': live_sessions,
        'all_sessions': all_sessions,
        'session_stats': session_stats,
        'recent_sessions': recent_sessions,
        
        'total_exams': total_exams,
        'active_exams': active_exams,
        'recent_exams': recent_exams,
        'total_attempts': total_attempts,
        
        'instructor_permissions': instructor_permissions,
    }


def get_student_statistics(user):
    """Get comprehensive statistics for student"""
    
    # Enrollments
    total_enrollments = Enrollment.objects.filter(
        student=user,
        is_active=True
    ).count()
    
    active_enrollments = Enrollment.objects.filter(
        student=user,
        status='enrolled',
        is_active=True
    ).count()
    
    completed_enrollments = Enrollment.objects.filter(
        student=user,
        status='completed',
        is_active=True
    ).count()
    
    # Get enrolled courses
    enrolled_courses = Enrollment.objects.filter(
        student=user,
        is_active=True
    ).select_related('course', 'course__instructor').order_by('-enrolled_at')
    
    # Batch Enrollments
    total_batch_enrollments = BatchEnrollment.objects.filter(
        student=user,
        is_active=True
    ).count()
    
    active_batch_enrollments = BatchEnrollment.objects.filter(
        student=user,
        status='enrolled',
        is_active=True
    ).count()
    
    enrolled_batches = BatchEnrollment.objects.filter(
        student=user,
        is_active=True
    ).select_related('batch', 'batch__course', 'batch__instructor').order_by('-enrolled_at')
    
    # Exam Attempts
    total_exam_attempts = ExamAttempt.objects.filter(
        student=user
    ).count()
    
    completed_exams = ExamAttempt.objects.filter(
        student=user,
        status__in=['submitted', 'auto_submitted']
    ).count()
    
    passed_exams = ExamAttempt.objects.filter(
        student=user,
        is_passed=True
    ).count()
    
    # Average percentage
    avg_percentage = ExamAttempt.objects.filter(
        student=user,
        status__in=['submitted', 'auto_submitted']
    ).aggregate(avg=Sum('percentage'))['avg'] or 0
    
    recent_exam_attempts = ExamAttempt.objects.filter(
        student=user
    ).select_related('exam').order_by('-started_at')[:5]
    
    # Session Attendance
    from zoom.models import SessionAttendance
    
    total_sessions_attended = SessionAttendance.objects.filter(
        student=user,
        is_present=True
    ).count()
    
    total_sessions_assigned = SessionAttendance.objects.filter(
        student=user
    ).count()
    
    attendance_percentage = 0
    if total_sessions_assigned > 0:
        attendance_percentage = round((total_sessions_attended / total_sessions_assigned) * 100, 2)
    
    recent_attendance = SessionAttendance.objects.filter(
        student=user
    ).select_related('session', 'session__batch').order_by('-created_at')[:5]
    
    return {
        'total_enrollments': total_enrollments,
        'active_enrollments': active_enrollments,
        'completed_enrollments': completed_enrollments,
        'enrolled_courses': enrolled_courses,
        
        'total_batch_enrollments': total_batch_enrollments,
        'active_batch_enrollments': active_batch_enrollments,
        'enrolled_batches': enrolled_batches,
        
        'total_exam_attempts': total_exam_attempts,
        'completed_exams': completed_exams,
        'passed_exams': passed_exams,
        'avg_percentage': round(avg_percentage, 2) if avg_percentage else 0,
        'recent_exam_attempts': recent_exam_attempts,
        
        'total_sessions_attended': total_sessions_attended,
        'total_sessions_assigned': total_sessions_assigned,
        'attendance_percentage': attendance_percentage,
        'recent_attendance': recent_attendance,
    }


# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 


# students/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Avg, Count, Sum
from django.utils import timezone
from django.http import JsonResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta

from userss.models import CustomUser, UserProfile, UserActivityLog
from courses.models import (
    Course, Enrollment, CourseCategory, 
    BatchEnrollment, LessonProgress
)
from .forms import StudentProfileForm ,UserProfileForm # You'll need to create this


# students/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Avg, Count, Sum
from django.utils import timezone
from django.http import JsonResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta



@login_required
def student_dashboard(request):
    """Enhanced student dashboard with profile completion check"""
    
    # Check if user is student
    if request.user.role != 'student':
        messages.error(request, "Access denied. Students only.")
        return redirect('user_login')
    
    # Profile completion check
    profile_complete = check_profile_completion(request.user)
    
    if not profile_complete:
        messages.warning(request, "Please complete your profile to access all features.")
        return redirect('student_profile')
    
    # Dashboard data
    user = request.user
    today = timezone.now().date()
    
    # Get user enrollments
    active_enrollments = Enrollment.objects.filter(
        student=user, 
        is_active=True
    ).select_related('course', 'course__instructor')
    
    # Get batch enrollments
    batch_enrollments = BatchEnrollment.objects.filter(
        student=user,
        is_active=True
    ).select_related('batch', 'batch__course')
    
    # Calculate stats
    total_courses = active_enrollments.count()
    total_batches = batch_enrollments.count()
    completed_courses = active_enrollments.filter(status='completed').count()
    
    # Overall progress calculation
    overall_progress = active_enrollments.aggregate(
        avg_progress=Avg('progress_percentage')
    )['avg_progress'] or 0
    
    # Recent activity
    recent_activities = UserActivityLog.objects.filter(
        user=user
    ).order_by('-timestamp')[:5]
    
    # Get some active courses for display
    active_courses = active_enrollments.order_by('-enrolled_at')[:3]
    
    # Performance data
    average_grade = active_enrollments.exclude(grade='').aggregate(
        avg=Avg('progress_percentage')
    )['avg'] or 0
    
    total_hours = active_enrollments.aggregate(
        total=Sum('total_time_spent_minutes')
    )['total'] or 0
    total_hours = round(total_hours / 60, 1)  # Convert to hours
    
    completion_rate = (completed_courses / total_courses * 100) if total_courses > 0 else 0
    
    # Upcoming deadlines (placeholder - you can extend this)
    upcoming_deadlines = []
    
    context = {
        'user': user,
        'today': today,
        'total_courses': total_courses,
        'total_batches': total_batches,
        'completed_courses': completed_courses,
        'overall_progress': overall_progress,
        'active_courses': active_courses,
        'recent_activities': recent_activities,
        'upcoming_deadlines': upcoming_deadlines,
        'average_grade': average_grade,
        'total_hours': total_hours,
        'completion_rate': completion_rate,
    }
    
    return render(request, 'student_dashboard.html', context)


def check_profile_completion(user):
    """Check if student profile is complete"""
    try:
        profile = user.profile
        # Check required fields
        required_fields = [
            profile.student_id,
            user.first_name,
            user.last_name,
            user.email,
        ]
        return all(field for field in required_fields)
    except:
        return False


@login_required
def student_profile(request):
    """Student profile completion/update view"""
    
    if request.user.role != 'student':
        messages.error(request, "Access denied.")
        return redirect('user_login')
    
    # Get or create profile
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)
    
    if request.method == 'POST':
        # Import forms here to avoid circular import
        from .forms import StudentProfileForm, UserProfileForm
        
        form = StudentProfileForm(request.POST, request.FILES, instance=request.user)
        profile_form = UserProfileForm(request.POST, instance=profile)
        
        if form.is_valid() and profile_form.is_valid():
            # Save user data
            user = form.save()
            
            # Save profile data
            profile = profile_form.save(commit=False)
            profile.user = user
            
            # Auto-generate student ID if not provided
            if not profile.student_id:
                profile.student_id = generate_student_id()
            
            profile.save()
            
            # Log activity
            UserActivityLog.objects.create(
                user=user,
                action='update_user',
                description='Profile updated successfully'
            )
            
            messages.success(request, 'Profile updated successfully!')
            return redirect('student_dashboard')
    else:
        # Import forms here to avoid circular import
        from .forms import StudentProfileForm, UserProfileForm
        
        form = StudentProfileForm(instance=request.user)
        profile_form = UserProfileForm(instance=profile)
    
    context = {
        'form': form,
        'profile_form': profile_form,
        'profile': profile,
        'is_profile_complete': check_profile_completion(request.user),
        'enrolled_courses_count': 0,  # Default value
        'batch_enrollments_count': 0,  # Default value
    }
    
    return render(request, 'students/student_profile.html', context)


def generate_student_id():
    """Generate unique student ID"""
    import random
    import string
    
    while True:
        # Generate format: STU + year + 4 random digits
        year = timezone.now().year
        random_part = ''.join(random.choices(string.digits, k=4))
        student_id = f"STU{year}{random_part}"
        
        # Check if unique
        if not UserProfile.objects.filter(student_id=student_id).exists():
            return student_id


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Prefetch
from courses.models import Enrollment, Course
from fees.models import StudentFeeAssignment, EMISchedule
from django.utils import timezone
from datetime import date

# courses/views.py - student_courses

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from courses.models import Enrollment,  BatchEnrollment
from fees.models import StudentFeeAssignment, EMISchedule

from fees.models import StudentFeeAssignment, EMISchedule

from fees.models import StudentFeeAssignment, EMISchedule

from fees.models import StudentFeeAssignment, EMISchedule

@login_required
def student_courses(request):
    """Student's enrolled courses - ONLY show enrolled courses"""
    
    if request.user.role != 'student':
        return redirect('user_login')
    
    from courses.models import Course, CourseReview  # ✅ CourseReview import add kiya
    from fees.models import StudentFeeAssignment, EMISchedule
    
    # ⭐ Get ONLY enrolled courses (not all courses)
    enrollments = Enrollment.objects.filter(
        student=request.user,
        is_active=True
    ).select_related(
        'course', 
        'course__instructor',
        'course__category'
    ).order_by('-enrolled_at')
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        enrollments = enrollments.filter(status=status)
    
    # Search
    search = request.GET.get('search')
    if search:
        enrollments = enrollments.filter(
            Q(course__title__icontains=search) |
            Q(course__course_code__icontains=search)
        )
    
    # Enhanced enrollment data
    enhanced_enrollments = []
    
    for enrollment in enrollments:
        course = enrollment.course
        
        # Fee assignment check
        try:
            fee_assignment = StudentFeeAssignment.objects.get(
                student=request.user,
                course=course
            )
        except StudentFeeAssignment.DoesNotExist:
            fee_assignment = None
        
        # ⭐ ENHANCED LOCK LOGIC - Status + Payment based
        lock_info = {
            'is_locked': False,
            'lock_reason': '',
            'warning_message': '',
            'can_unlock': False,
            'action_required': '',
            'overdue_amount': 0,
        }
        
        # 🔒 PRIORITY 1: Check enrollment status first
        if enrollment.status == 'dropped':
            lock_info['is_locked'] = True
            lock_info['lock_reason'] = '🚫 Course Dropped. You have withdrawn from this course.'
            lock_info['can_unlock'] = False
            lock_info['action_required'] = 'contact_admin'
            
        elif enrollment.status == 'suspended':
            lock_info['is_locked'] = True
            lock_info['lock_reason'] = '⏸️ Course Suspended. Your access has been temporarily suspended.'
            lock_info['can_unlock'] = False
            lock_info['action_required'] = 'contact_admin'
        
        # 🔒 PRIORITY 2: Check payment issues (only if not dropped/suspended)
        elif enrollment.status == 'enrolled':
            if fee_assignment:
                if fee_assignment.amount_pending > 0:
                    # Check if course is locked due to payment
                    if hasattr(fee_assignment, 'is_course_locked') and fee_assignment.is_course_locked:
                        lock_info['is_locked'] = True
                        lock_info['lock_reason'] = '💳 Payment Pending. Complete your fees to unlock access.'
                        lock_info['can_unlock'] = True
                        lock_info['action_required'] = 'make_payment'
                        lock_info['overdue_amount'] = float(fee_assignment.amount_pending)
        
        # ✅ Show completion status
        elif enrollment.status == 'completed':
            lock_info['is_locked'] = False
            lock_info['warning_message'] = '✅ Course Completed - View Only Mode'
        
        # Get next due payment (warning only - if not already locked)
        next_due = None
        if fee_assignment and fee_assignment.fee_structure.payment_type == 'emi' and not lock_info['is_locked']:
            next_due = EMISchedule.objects.filter(
                fee_assignment=fee_assignment,
                status__in=['pending', 'overdue']
            ).order_by('due_date').first()
            
            if next_due:
                from datetime import date
                days_until_due = (next_due.due_date - date.today()).days
                if days_until_due <= 7 and days_until_due >= 0:
                    lock_info['warning_message'] = f'⚠️ Payment of ₹{next_due.amount} due in {days_until_due} days'
                elif days_until_due < 0:
                    lock_info['warning_message'] = f'🔴 Payment of ₹{next_due.amount} overdue by {abs(days_until_due)} days!'
        
        # ✅ CHECK IF STUDENT HAS REVIEWED THIS COURSE
        has_reviewed = CourseReview.objects.filter(
            student=request.user,
            course=course
        ).exists()
        
        enhanced_enrollments.append({
            'enrollment': enrollment,
            'course': course,
            'fee_assignment': fee_assignment,
            'lock_info': lock_info,
            'next_due': next_due,
            'has_reviewed': has_reviewed,  # ✅ YEH ADD KIYA
        })
    
    # Pagination
    paginator = Paginator(enhanced_enrollments, 12)
    page = request.GET.get('page')
    enhanced_enrollments = paginator.get_page(page)
    
    # Summary statistics
    total_courses = enrollments.count()
    locked_courses = sum(1 for item in enhanced_enrollments.object_list if item['lock_info']['is_locked'])
    dropped_courses = enrollments.filter(status='dropped').count()
    suspended_courses = enrollments.filter(status='suspended').count()
    completed_courses = enrollments.filter(status='completed').count()
    
    context = {
        'enrollments': enhanced_enrollments,
        'status': status,
        'search': search,
        'total_courses': total_courses,
        'locked_courses': locked_courses,
        'active_courses': enrollments.filter(status='enrolled').count(),
        'dropped_courses': dropped_courses,
        'suspended_courses': suspended_courses,
        'completed_courses': completed_courses,
    }
    
    return render(request, 'students/student_courses.html', context)

def get_course_lock_info(course, student, fee_assignment=None):
    """
    Get comprehensive course lock information
    Returns dict with lock status, reason, and related info
    """
    lock_info = {
        'is_locked': False,
        'lock_reason': None,
        'can_unlock': False,
        'unlock_date': None,
        'payment_status': 'current',
        'overdue_amount': 0,
        'overdue_days': 0,
        'grace_period_remaining': 0,
        'warning_message': None,
        'action_required': None,
    }
    
    if not fee_assignment:
        # No fee assignment means course is accessible
        return lock_info
    
    # Check if course is manually locked
    if fee_assignment.is_course_locked:
        lock_info.update({
            'is_locked': True,
            'lock_reason': 'Payment overdue',
            'can_unlock': False,
        })
        
        # Check if admin set unlock date
        if fee_assignment.unlock_date:
            lock_info.update({
                'unlock_date': fee_assignment.unlock_date,
                'can_unlock': date.today() >= fee_assignment.unlock_date,
            })
    
    # Check payment status for EMI plans
    if fee_assignment.fee_structure.payment_type == 'emi':
        overdue_emis = EMISchedule.objects.filter(
            fee_assignment=fee_assignment,
            status__in=['pending', 'overdue'],
            due_date__lt=date.today()
        ).order_by('due_date')
        
        if overdue_emis.exists():
            oldest_overdue = overdue_emis.first()
            overdue_days = (date.today() - oldest_overdue.due_date).days
            overdue_amount = sum(emi.amount for emi in overdue_emis)
            
            lock_info.update({
                'payment_status': 'overdue',
                'overdue_amount': overdue_amount,
                'overdue_days': overdue_days,
            })
            
            # Check grace period
            grace_period = fee_assignment.fee_structure.grace_period_days
            if overdue_days <= grace_period:
                lock_info.update({
                    'grace_period_remaining': grace_period - overdue_days,
                    'warning_message': f'Payment overdue by {overdue_days} days. Grace period expires in {grace_period - overdue_days} days.',
                    'action_required': 'pay_now'
                })
            else:
                # Should be locked
                if not fee_assignment.is_course_locked:
                    # Auto-lock the course
                    fee_assignment.lock_course()
                
                lock_info.update({
                    'is_locked': True,
                    'lock_reason': f'Payment overdue by {overdue_days} days (grace period exceeded)',
                    'action_required': 'contact_admin'
                })
    
    # Check for upcoming due dates (warning)
    if not lock_info['is_locked'] and fee_assignment.fee_structure.payment_type == 'emi':
        upcoming_emi = EMISchedule.objects.filter(
            fee_assignment=fee_assignment,
            status='pending',
            due_date__gte=date.today(),
            due_date__lte=date.today() + timezone.timedelta(days=7)
        ).order_by('due_date').first()
        
        if upcoming_emi:
            days_until_due = (upcoming_emi.due_date - date.today()).days
            lock_info.update({
                'warning_message': f'Next payment of ${upcoming_emi.amount} due in {days_until_due} days',
                'action_required': 'payment_due_soon'
            })
    
    return lock_info


def get_payment_summary(student, course):
    """Get payment summary for a student's course"""
    try:
        fee_assignment = StudentFeeAssignment.objects.get(
            student=student,
            course=course
        )
        
        summary = {
            'total_amount': fee_assignment.total_amount,
            'amount_paid': fee_assignment.amount_paid,
            'amount_pending': fee_assignment.amount_pending,
            'completion_percentage': fee_assignment.get_completion_percentage(),
            'payment_type': fee_assignment.fee_structure.payment_type,
        }
        
        # EMI specific info
        if fee_assignment.fee_structure.payment_type == 'emi':
            total_emis = fee_assignment.emi_schedules.count()
            paid_emis = fee_assignment.emi_schedules.filter(status='paid').count()
            
            summary.update({
                'total_emis': total_emis,
                'paid_emis': paid_emis,
                'remaining_emis': total_emis - paid_emis,
            })
        
        return summary
        
    except StudentFeeAssignment.DoesNotExist:
        return None


# Helper view for AJAX requests to check lock status
@login_required
def check_course_lock_status(request, course_id):
    """AJAX endpoint to check course lock status"""
    
    if request.user.role != 'student':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        course = Course.objects.get(id=course_id)
        fee_assignment = StudentFeeAssignment.objects.get(
            student=request.user,
            course=course
        )
        
        lock_info = get_course_lock_info(course, request.user, fee_assignment)
        
        return JsonResponse({
            'success': True,
            'lock_info': lock_info
        })
        
    except Course.DoesNotExist:
        return JsonResponse({'error': 'Course not found'}, status=404)
    except StudentFeeAssignment.DoesNotExist:
        return JsonResponse({'error': 'Fee assignment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    


@login_required
def student_batches(request):
    """Student's batch enrollments - DEBUG VERSION"""
    
    if request.user.role != 'student':
        return redirect('user_login')
    
    from fees.models import StudentFeeAssignment
    from django.db.models import Q
    
    # Get ALL enrollments
    batch_enrollments = BatchEnrollment.objects.filter(
        student=request.user
    ).select_related('batch', 'batch__course', 'batch__instructor')
    
    print("\n" + "="*60)
    print(f"DEBUGGING FOR USER: {request.user.username}")
    print(f"Total enrollments found: {batch_enrollments.count()}")
    print("="*60)
    
    # Filters
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    sort = request.GET.get('sort', '-enrolled_at')
    
    if search:
        batch_enrollments = batch_enrollments.filter(
            Q(batch__name__icontains=search) |
            Q(batch__course__title__icontains=search) |
            Q(batch__code__icontains=search)
        )
    
    if sort:
        batch_enrollments = batch_enrollments.order_by(sort)
    
    # Stats
    active_batches_count = 0
    locked_batches_count = 0
    completed_batches_count = 0
    upcoming_batches_count = 0
    
    filtered_enrollments = []
    
    for enrollment in batch_enrollments:
        print(f"\n--- Enrollment ID: {enrollment.id} ---")
        print(f"Batch: {enrollment.batch.name}")
        print(f"Batch Status: {enrollment.batch.status}")
        print(f"Enrollment is_active: {enrollment.is_active}")
        
        # Fee assignment
        try:
            enrollment.fee_assignment = StudentFeeAssignment.objects.get(
                student=request.user,
                course=enrollment.batch.course
            )
            print(f"Fee Assignment found: {enrollment.fee_assignment}")
            print(f"Fee is_course_locked: {enrollment.fee_assignment.is_course_locked}")
        except StudentFeeAssignment.DoesNotExist:
            enrollment.fee_assignment = None
            print(f"Fee Assignment: None")
        
        # Check locked status
        print(f"enrollment.is_locked property: {enrollment.is_locked}")
        print(f"enrollment.get_lock_reason(): {enrollment.get_lock_reason()}")
        
        batch_status = enrollment.batch.status
        
        # Determine if locked
        is_enrollment_locked = enrollment.is_locked
        
        print(f"Final determination - is_enrollment_locked: {is_enrollment_locked}")
        
        if batch_status == 'completed':
            completed_batches_count += 1
            print(f"Categorized as: COMPLETED")
            if status_filter == '' or status_filter == 'completed':
                filtered_enrollments.append(enrollment)
                
        elif batch_status == 'draft':
            upcoming_batches_count += 1
            print(f"Categorized as: UPCOMING")
            if status_filter == '' or status_filter == 'upcoming':
                filtered_enrollments.append(enrollment)
                
        elif batch_status == 'active':
            if is_enrollment_locked:
                locked_batches_count += 1
                print(f"Categorized as: LOCKED ✓")
                if status_filter == '' or status_filter == 'locked':
                    filtered_enrollments.append(enrollment)
            else:
                active_batches_count += 1
                print(f"Categorized as: ACTIVE ✓")
                if status_filter == '' or status_filter == 'active':
                    filtered_enrollments.append(enrollment)
        else:
            print(f"Categorized as: OTHER")
            if status_filter == '':
                filtered_enrollments.append(enrollment)
    
    print("\n" + "="*60)
    print("FINAL COUNTS:")
    print(f"Active: {active_batches_count}")
    print(f"Locked: {locked_batches_count}")
    print(f"Completed: {completed_batches_count}")
    print(f"Upcoming: {upcoming_batches_count}")
    print(f"Filtered enrollments: {len(filtered_enrollments)}")
    print("="*60 + "\n")
    
    context = {
        'batch_enrollments': filtered_enrollments,
        'active_batches_count': active_batches_count,
        'locked_batches_count': locked_batches_count,
        'completed_batches_count': completed_batches_count,
        'upcoming_batches_count': upcoming_batches_count,
        'search': search,
        'status': status_filter,
        'sort': sort,
    }
    
    return render(request, 'students/student_batches.html', context)


@login_required
def browse_courses(request):
    """Browse available courses - Check enrollment status for locks"""
    
    if request.user.role != 'student':
        return redirect('user_login')
    
    from courses.models import Enrollment, Course, CourseCategory
    from fees.models import StudentFeeAssignment
    from django.db.models import Q
    from django.core.paginator import Paginator
    from collections import OrderedDict
    
    # Get courses
    courses = Course.objects.filter(
        is_active=True,
        status='published'
    ).select_related('instructor', 'category')
    
    # Filters
    category = request.GET.get('category')
    if category:
        courses = courses.filter(category__slug=category)
    
    difficulty = request.GET.get('difficulty')
    if difficulty:
        courses = courses.filter(difficulty_level=difficulty)
    
    course_type = request.GET.get('type')
    if course_type:
        courses = courses.filter(course_type=course_type)
    
    search = request.GET.get('search')
    if search:
        courses = courses.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(course_code__icontains=search)
        )
    
    sort = request.GET.get('sort', '-created_at')
    courses = courses.order_by(sort)
    
    # ⭐⭐⭐ ENHANCED CHECK: Course Enrollment + Status ⭐⭐⭐
    for course in courses:
        # ⭐ Check direct course enrollment
        course_enrollment = Enrollment.objects.filter(
            student=request.user,
            course=course,
            is_active=True
        ).first()
        
        # Default values
        course.is_student_enrolled = bool(course_enrollment)
        course.is_course_locked = False
        course.lock_reason = ''
        course.enrollment_status = None
        
        # 🔒 Check enrollment status and payment
        if course_enrollment:
            course.enrollment_status = course_enrollment.status
            
            # 🔒 PRIORITY 1: Check if dropped or suspended
            if course_enrollment.status == 'dropped':
                course.is_course_locked = True
                course.lock_reason = '🚫 Course Dropped - You have withdrawn from this course'
                
            elif course_enrollment.status == 'suspended':
                course.is_course_locked = True
                course.lock_reason = '⏸️ Access Suspended - Please contact administration'
            
            # 🔒 PRIORITY 2: Check payment (only if enrolled status)
            elif course_enrollment.status == 'enrolled':
                try:
                    fee_assignment = StudentFeeAssignment.objects.get(
                        student=request.user,
                        course=course
                    )
                    # Check if course is locked due to payment
                    if hasattr(fee_assignment, 'is_course_locked') and fee_assignment.is_course_locked:
                        course.is_course_locked = True
                        course.lock_reason = f'💳 Payment Pending - ₹{fee_assignment.amount_pending} due'
                except StudentFeeAssignment.DoesNotExist:
                    pass
            
            # ✅ Completed courses are accessible (view only)
            elif course_enrollment.status == 'completed':
                course.is_course_locked = False
                course.lock_reason = '✅ Course Completed'
        
        # Check if enrollment is open (for non-enrolled students)
        course.is_enrollment_open_flag = course.is_enrollment_open()
        course.available_seats = course.get_available_seats()
    
    # Group by category
    category_courses = OrderedDict()
    
    for course in courses:
        cat = course.category
        if cat not in category_courses:
            category_courses[cat] = []
        category_courses[cat].append(course)
    
    # Pagination
    paginator = Paginator(courses, 12)
    page = request.GET.get('page')
    courses_page = paginator.get_page(page)
    
    categories = CourseCategory.objects.filter(is_active=True)
    
    context = {
        'courses': courses_page,
        'categories': categories,
        'selected_category': category,
        'selected_difficulty': difficulty,
        'selected_type': course_type,
        'search': search,
        'sort': sort,
        'total_courses': Course.objects.filter(is_active=True, status='published').count(),
        'category_courses': category_courses,
        'total_categories': len(category_courses),
    }
    
    return render(request, 'students/browse_courses.html', context)


@login_required
def student_progress(request):
    """Student progress tracking"""
    
    if request.user.role != 'student':
        return redirect('user_login')
    
    # Get all enrollments with progress
    enrollments = Enrollment.objects.filter(
        student=request.user,
        is_active=True
    ).select_related('course')
    
    # Get detailed progress for each course
    progress_data = []
    for enrollment in enrollments:
        # Get lesson progress for this course
        lesson_progress = LessonProgress.objects.filter(
            student=request.user,
            lesson__module__course=enrollment.course
        )
        
        total_lessons = enrollment.course.modules.filter(
            is_active=True
        ).aggregate(
            total=Count('lessons', filter=Q(lessons__is_active=True))
        )['total'] or 0
        
        completed_lessons = lesson_progress.filter(
            status='completed'
        ).count()
        
        progress_data.append({
            'enrollment': enrollment,
            'total_lessons': total_lessons,
            'completed_lessons': completed_lessons,
            'completion_rate': (completed_lessons / total_lessons * 100) if total_lessons > 0 else 0
        })
    
    context = {
        'progress_data': progress_data,
    }
    
    return render(request, 'students/student_progress.html', context)


@login_required
def student_assignments(request):
    """Student assignments (placeholder)"""
    
    if request.user.role != 'student':
        return redirect('user_login')
    
    # This is a placeholder - you can extend with actual assignment models
    context = {
        'assignments': [],
    }
    
    return render(request, 'students/student_assignments.html', context)


@login_required
def student_certificates(request):
    """Student certificates (placeholder)"""
    
    if request.user.role != 'student':
        return redirect('user_login')
    
    # Get completed courses
    completed_enrollments = Enrollment.objects.filter(
        student=request.user,
        status='completed',
        is_active=True
    ).select_related('course')
    
    context = {
        'completed_enrollments': completed_enrollments,
    }
    
    return render(request, 'students/student_certificates.html', context)


@login_required
def student_settings(request):
    """Student settings"""
    
    if request.user.role != 'student':
        return redirect('user_login')
    
    if request.method == 'POST':
        # Handle settings update
        pass
    
    context = {}
    
    return render(request, 'students/student_settings.html', context)



# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# student dashborad views 
# views.py - Add these views to your existing views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from django.db.models import Q, Count, Sum
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.core.paginator import Paginator


import logging
logger = logging.getLogger(__name__)

def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# Your existing views (keep as is, just add email sending functionality)

# Add these views to your existing views.py

from .models import EmailTemplate, EmailLog, DailyEmailSummary, EmailLimitSet
from django.db.models import Count
from datetime import date, timedelta



@login_required
def manage_email_templates(request):
    """Manage email templates"""
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    templates = EmailTemplate.objects.all().order_by('template_type')
    return render(request, 'email_management/templates.html', {'templates': templates})


@login_required
def edit_email_template(request, template_id):
    """Edit email template"""
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    template = get_object_or_404(EmailTemplate, id=template_id)
    
    if request.method == 'POST':
        template.name = request.POST.get('name')
        template.subject = request.POST.get('subject')
        template.email_body = request.POST.get('email_body')
        template.is_active = 'is_active' in request.POST
        template.save()
        
        messages.success(request, 'Email template updated successfully!')
        return redirect('manage_email_templates')
    
    return render(request, 'email_management/edit_template.html', {'template': template})


@login_required
def email_limit_settings(request):
    """Manage email limit settings"""
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    email_limit_setting = EmailLimitSet.objects.filter(is_active=True).first()
    
    if request.method == 'POST':
        new_limit = request.POST.get('email_limit_per_day')
        try:
            new_limit = int(new_limit)
            if new_limit > 0:
                # Deactivate old settings
                EmailLimitSet.objects.filter(is_active=True).update(is_active=False)
                
                # Create new setting
                EmailLimitSet.objects.create(
                    email_limit_per_day=new_limit,
                    is_active=True
                )
                
                messages.success(request, f'Daily email limit updated to {new_limit}!')
            else:
                messages.error(request, 'Email limit must be greater than 0!')
        except ValueError:
            messages.error(request, 'Please enter a valid number!')
        
        return redirect('email_limit_settings')
    
    return render(request, 'email_management/limit_settings.html', {
        'email_limit_setting': email_limit_setting
    })

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import EmailSMTPConfiguration

@login_required
def email_smtp_settings(request):
    """Manage Email SMTP Configuration (NOT limits - that's separate)"""
    if request.user.role != 'superadmin':
        messages.error(request, 'Only Super Admin can access this page!')
        return redirect('user_login')
    
    # Get active SMTP configuration
    smtp_config = EmailSMTPConfiguration.objects.filter(is_active=True).first()
    
    if request.method == 'POST':
        # Extract form data
        email_host = request.POST.get('email_host', 'smtp.gmail.com').strip()
        email_port = request.POST.get('email_port', 587)
        email_use_tls = request.POST.get('email_use_tls') == 'on'
        email_host_user = request.POST.get('email_host_user', '').strip()
        email_host_password = request.POST.get('email_host_password', '').strip()
        default_from_email = request.POST.get('default_from_email', '').strip()
        
        try:
            # Validate data
            email_port = int(email_port)
            
            if not email_host_user or not email_host_password:
                messages.error(request, 'Email and Password are required!')
                return redirect('email_smtp_settings')
            
            # Deactivate old configurations
            EmailSMTPConfiguration.objects.filter(is_active=True).update(is_active=False)
            
            # Create new configuration
            EmailSMTPConfiguration.objects.create(
                email_host=email_host,
                email_port=email_port,
                email_use_tls=email_use_tls,
                email_host_user=email_host_user,
                email_host_password=email_host_password,
                default_from_email=default_from_email or email_host_user,
                is_active=True
            )
            
            messages.success(request, '✅ Email SMTP configuration updated successfully!')
            return redirect('email_smtp_settings')
            
        except ValueError:
            messages.error(request, 'Please enter valid port number!')
            return redirect('email_smtp_settings')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('email_smtp_settings')
    
    return render(request, 'email_configuration.html', {
        'smtp_config': smtp_config
    })

@login_required
def email_logs(request):
    """View email logs with filtering"""
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    # Filters
    date_filter = request.GET.get('date', '')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    logs = EmailLog.objects.all()
    
    # Apply filters
    if date_filter:
        logs = logs.filter(sent_date=date_filter)
    if status_filter:
        if status_filter == 'success':
            logs = logs.filter(is_sent_successfully=True)
        elif status_filter == 'failed':
            logs = logs.filter(is_sent_successfully=False)
    if search_query:
        logs = logs.filter(
            Q(recipient_email__icontains=search_query) |
            Q(subject__icontains=search_query)
        )
    
    # Calculate statistics before pagination
    total_count = logs.count()
    successful_count = logs.filter(is_sent_successfully=True).count()
    failed_count = logs.filter(is_sent_successfully=False).count()
    success_rate = round((successful_count / total_count * 100), 1) if total_count > 0 else 0
    
    logs = logs.order_by('-sent_time')
    
    # Pagination
    paginator = Paginator(logs, 20)
    page = request.GET.get('page')
    logs = paginator.get_page(page)
    
    context = {
        'logs': logs,
        'date_filter': date_filter,
        'status_filter': status_filter,
        'search_query': search_query,
        'successful_count': successful_count,
        'failed_count': failed_count,
        'success_rate': success_rate,
    }
    return render(request, 'email_management/logs.html', context)

# Add these views to your existing views.py

from django.core.mail import send_mass_mail
from django.db import transaction
from django.http import JsonResponse
import json


import re
from datetime import date
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
import logging

logger = logging.getLogger(__name__)

def replace_template_variables(text, user):
    """
    Template se automatically variables detect karo aur user data se replace karo
    """
    # Pehle template se saare {{variable}} dhundo
    pattern = r'\{\{(\w+)\}\}'
    variables = re.findall(pattern, text)
    
    # Ab har variable ko user ke data se replace karo
    for var in variables:
        var_lower = var.lower().strip()
        value = ''
        
        # Variable type check karo aur corresponding data lo
        if var_lower in ['username', 'user']:
            value = user.username if user.username else user.email.split('@')[0]
        
        elif var_lower in ['email', 'mail']:
            value = user.email if user.email else ''
        
        elif var_lower in ['first_name', 'firstname', 'fname']:
            value = user.first_name if user.first_name else ''
        
        elif var_lower in ['last_name', 'lastname', 'lname']:
            value = user.last_name if user.last_name else ''
        
        elif var_lower in ['full_name', 'fullname', 'name']:
            if user.first_name and user.last_name:
                value = f"{user.first_name} {user.last_name}"
            elif user.first_name:
                value = user.first_name
            elif user.last_name:
                value = user.last_name
            elif user.username:
                value = user.username
            else:
                value = user.email.split('@')[0]
        
        # Replace karo
        text = text.replace('{{' + var + '}}', value)
    
    return text


@login_required
def bulk_email(request):
    """Send bulk emails to selected users"""
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    if request.method == 'POST':
        template_id = request.POST.get('template_id')
        recipient_type = request.POST.get('recipient_type')
        custom_subject = request.POST.get('custom_subject')
        custom_message = request.POST.get('custom_message')
        selected_users = request.POST.getlist('selected_users')
        role = request.POST.get('role')
        
        if not custom_subject or not custom_message:
            messages.error(request, 'Subject and message are required!')
            return redirect('bulk_email')
        
        if check_daily_email_limit():
            messages.error(request, 'Daily email limit reached! Cannot send bulk emails.')
            return redirect('bulk_email')
        
        recipients = []
        if recipient_type == 'all':
            recipients = CustomUser.objects.filter(email__isnull=False).exclude(email='')
        elif recipient_type == 'role':
            if not role:
                messages.error(request, 'Please select a role!')
                return redirect('bulk_email')
            recipients = CustomUser.objects.filter(role=role, email__isnull=False).exclude(email='')
        elif recipient_type == 'selected':
            if not selected_users:
                messages.error(request, 'Please select at least one user!')
                return redirect('bulk_email')
            recipients = CustomUser.objects.filter(id__in=selected_users, email__isnull=False).exclude(email='')
        
        if not recipients.exists():
            messages.error(request, 'No valid recipients found!')
            return redirect('bulk_email')
        
        template = None
        if template_id:
            try:
                template = EmailTemplate.objects.get(id=template_id, is_active=True)
            except EmailTemplate.DoesNotExist:
                pass
        
        successful_count = 0
        failed_count = 0
        failed_emails = []
        
        for user in recipients:
            if check_daily_email_limit():
                messages.warning(request, f'Daily email limit reached! Only {successful_count} emails were sent.')
                break
            
            try:
                if template:
                    subject = template.subject
                    message = template.email_body
                else:
                    subject = custom_subject
                    message = custom_message
                
                # MAIN MAGIC: Template se automatically variables detect aur replace
                subject = replace_template_variables(subject, user)
                message = replace_template_variables(message, user)
                
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email],
                    fail_silently=False
                )
                
                successful_count += 1
                
                EmailLog.objects.create(
                    recipient_email=user.email,
                    recipient_user=user,
                    template_used=template,
                    subject=subject,
                    email_body=message,
                    is_sent_successfully=True,
                    sent_by=request.user
                )
                
            except Exception as e:
                failed_count += 1
                failed_emails.append(f"{user.email}: {str(e)}")
                
                EmailLog.objects.create(
                    recipient_email=user.email,
                    recipient_user=user,
                    template_used=template,
                    subject=custom_subject,
                    email_body=custom_message,
                    is_sent_successfully=False,
                    error_message=str(e),
                    sent_by=request.user
                )
        
        try:
            today = date.today()
            email_limit = 50
            if EmailLimitSet.objects.filter(is_active=True).exists():
                email_limit = EmailLimitSet.objects.filter(is_active=True).first().email_limit_per_day
            
            daily_summary, created = DailyEmailSummary.objects.get_or_create(
                date=today,
                defaults={'daily_limit': email_limit}
            )
            daily_summary.total_emails_sent += (successful_count + failed_count)
            daily_summary.successful_emails += successful_count
            daily_summary.failed_emails += failed_count
            daily_summary.save()
        except Exception as e:
            logger.error(f'Error updating daily summary: {e}')
        
        if successful_count > 0:
            messages.success(request, f'Bulk email sent successfully to {successful_count} recipients!')
        
        if failed_count > 0:
            messages.warning(request, f'{failed_count} emails failed to send. Check email logs for details.')
            logger.warning(f'Failed emails: {failed_emails}')
        
        if successful_count == 0 and failed_count == 0:
            messages.error(request, 'No emails were sent!')
        
        return redirect('email_dashboard')
    
    # GET request
    users = CustomUser.objects.filter(
        email__isnull=False
    ).exclude(
        email=''
    ).order_by('first_name', 'username')
    
    templates = EmailTemplate.objects.filter(is_active=True).order_by('name')
    
    students_count = users.filter(role='student').count()
    instructors_count = users.filter(role='instructor').count()
    admins_count = users.filter(role='superadmin').count()
    
    context = {
        'users': users,
        'templates': templates,
        'students_count': students_count,
        'instructors_count': instructors_count,
        'admins_count': admins_count,
        'total_users': users.count(),
    }
    
    return render(request, 'email_management/bulk_email.html', context)


@login_required
def create_email_template(request):
    """Create new email template"""
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        template_type = request.POST.get('template_type')
        subject = request.POST.get('subject')
        email_body = request.POST.get('email_body')
        is_active = 'is_active' in request.POST
        
        # Check if template type already exists
        if EmailTemplate.objects.filter(template_type=template_type).exists():
            messages.error(request, 'Template type already exists!')
            return redirect('create_email_template')
        
        EmailTemplate.objects.create(
            name=name,
            template_type=template_type,
            subject=subject,
            email_body=email_body,
            is_active=is_active
        )
        
        messages.success(request, 'Email template created successfully!')
        return redirect('manage_email_templates')
    
    return render(request, 'email_management/create_template.html', {
        'template_types': EmailTemplate.TEMPLATE_TYPES
    })


@login_required
def email_dashboard(request):
    """Enhanced email management dashboard"""
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    # Get email statistics
    today = date.today()
    yesterday = today - timedelta(days=1)
    
    today_summary = DailyEmailSummary.objects.filter(date=today).first()
    yesterday_summary = DailyEmailSummary.objects.filter(date=yesterday).first()
    
    # Get current email limit
    email_limit_setting = EmailLimitSet.objects.filter(is_active=True).first()
    
    # Recent email logs
    recent_emails = EmailLog.objects.order_by('-sent_time')[:10]
    
    # Email statistics for last 7 days
    week_ago = today - timedelta(days=7)
    weekly_stats = DailyEmailSummary.objects.filter(
        date__gte=week_ago
    ).order_by('-date')
    
    # Get email templates
    email_templates = EmailTemplate.objects.filter(is_active=True)[:5]
    
    context = {
        'today_summary': today_summary,
        'yesterday_summary': yesterday_summary,
        'email_limit_setting': email_limit_setting,
        'recent_emails': recent_emails,
        'weekly_stats': weekly_stats,
        'email_templates': email_templates,
    }
    return render(request, 'email_management/dashboard.html', context)


@login_required
def email_analytics(request):
    """Email analytics and reports"""
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    from django.db.models import Count, Sum
    from django.db.models.functions import TruncMonth, ExtractMonth
    from datetime import datetime
    
    # Get monthly data for current year
    current_year = datetime.now().year
    
    # Fix for SQLite - Use TruncMonth instead of extra()
    monthly_stats = DailyEmailSummary.objects.filter(
        date__year=current_year
    ).annotate(
        month=ExtractMonth('date')
    ).values('month').annotate(
        total_emails=Sum('total_emails_sent'),
        successful_emails=Sum('successful_emails'),
        failed_emails=Sum('failed_emails')
    ).order_by('month')
    
    # Template usage statistics
    template_stats = EmailLog.objects.values(
        'template_used__name'
    ).annotate(
        usage_count=Count('id')
    ).order_by('-usage_count')[:10]
    
    # Success rate by template
    from django.db.models import Q
    success_rates = EmailLog.objects.values(
        'template_used__name'
    ).annotate(
        total_sent=Count('id'),
        successful=Count('id', filter=Q(is_sent_successfully=True))
    ).order_by('-total_sent')
    
    context = {
        'monthly_stats': monthly_stats,
        'template_stats': template_stats,
        'success_rates': success_rates,
        'current_year': current_year,
    }
    return render(request, 'email_management/analytics.html', context)


@login_required
def test_email_template(request, template_id):
    """Test email template by sending to admin"""
    if request.user.role != 'superadmin':
        return JsonResponse({'success': False, 'message': 'Permission denied'})
    
    template = get_object_or_404(EmailTemplate, id=template_id)
    
    # Check daily email limit
    if check_daily_email_limit():
        return JsonResponse({'success': False, 'message': 'Daily email limit reached'})
    
    try:
        # Replace template variables with admin data
        subject = template.subject
        message = template.email_body.replace(
            '{{username}}', request.user.username
        ).replace(
            '{{email}}', request.user.email
        ).replace(
            '{{first_name}}', request.user.first_name or request.user.username
        ).replace(
            '{{last_name}}', request.user.last_name or ''
        )
        
        # Send test email
        send_mail(
            subject=f"[TEST] {subject}",
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[request.user.email],
            fail_silently=False,
        )
        
        # Log test email
        EmailLog.objects.create(
            recipient_email=request.user.email,
            recipient_user=request.user,
            template_used=template,
            subject=f"[TEST] {subject}",
            email_body=message,
            is_sent_successfully=True,
            sent_by=request.user
        )
        
        return JsonResponse({'success': True, 'message': 'Test email sent successfully!'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Failed to send test email: {str(e)}'})


@login_required
def delete_email_template(request, template_id):
    """Delete email template"""
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    template = get_object_or_404(EmailTemplate, id=template_id)
    
    if request.method == 'POST':
        template_name = template.name
        template.delete()
        messages.success(request, f'Email template "{template_name}" deleted successfully!')
        return redirect('manage_email_templates')
    
    return render(request, 'email_management/delete_template.html', {'template': template})


# Add this missing function to your views.py

def check_daily_email_limit():
    """Check if daily email limit is reached"""
    try:
        email_limit_setting = EmailLimitSet.objects.filter(is_active=True).first()
        if not email_limit_setting:
            return False  # No limit set, can send
        
        today = date.today()
        daily_summary, created = DailyEmailSummary.objects.get_or_create(
            date=today,
            defaults={'daily_limit': email_limit_setting.email_limit_per_day}
        )
        
        return daily_summary.total_emails_sent >= daily_summary.daily_limit
    except Exception:
        return False
    

# Add these views to your views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
from django.db import IntegrityError
from django.core.paginator import Paginator
from .models import EmailTemplateType, EmailTemplate

@login_required
@staff_member_required
def manage_template_types(request):
    """Display all template types with statistics"""
    template_types = EmailTemplateType.objects.all().order_by('-created_at')
    
    # Statistics
    total_types = template_types.count()
    active_types = template_types.filter(is_active=True).count()
    inactive_types = template_types.filter(is_active=False).count()
    templates_count = EmailTemplate.objects.count()
    
    # Add template count for each type
    for template_type in template_types:
        template_type.template_count = template_type.templates.count()
    
    context = {
        'template_types': template_types,
        'total_types': total_types,
        'active_types': active_types,
        'inactive_types': inactive_types,
        'templates_count': templates_count,
    }
    
    return render(request, 'email_management/manage_template_types.html', context)


@login_required
@staff_member_required
def create_template_type(request):
    """Create new template type"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            code = request.POST.get('code', '').strip().lower()
            description = request.POST.get('description', '').strip()
            is_active = request.POST.get('is_active') == 'on'
            
            # Validation
            if not name or not code:
                messages.error(request, 'Name and Code are required fields!')
                return redirect('manage_template_types')
            
            # Check if code contains only valid characters
            if not code.replace('_', '').replace('-', '').isalnum():
                messages.error(request, 'Code can only contain letters, numbers, underscores and hyphens!')
                return redirect('manage_template_types')
            
            # Create template type
            template_type = EmailTemplateType.objects.create(
                name=name,
                code=code,
                description=description,
                is_active=is_active,
                created_by=request.user
            )
            
            messages.success(
                request, 
                f'Template type "{template_type.name}" has been created successfully!'
            )
            
        except IntegrityError as e:
            if 'name' in str(e):
                messages.error(request, 'A template type with this name already exists!')
            elif 'code' in str(e):
                messages.error(request, 'A template type with this code already exists!')
            else:
                messages.error(request, 'An error occurred while creating the template type.')
        
        except Exception as e:
            messages.error(request, f'An unexpected error occurred: {str(e)}')
    
    return redirect('manage_template_types')


@login_required
@staff_member_required
def update_template_type(request, template_type_id):
    """Update existing template type"""
    template_type = get_object_or_404(EmailTemplateType, id=template_type_id)
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            code = request.POST.get('code', '').strip().lower()
            description = request.POST.get('description', '').strip()
            is_active = request.POST.get('is_active') == 'on'
            
            # Validation
            if not name or not code:
                messages.error(request, 'Name and Code are required fields!')
                return redirect('manage_template_types')
            
            # Check for duplicate name (excluding current record)
            if EmailTemplateType.objects.filter(name=name).exclude(id=template_type_id).exists():
                messages.error(request, 'A template type with this name already exists!')
                return redirect('manage_template_types')
            
            # Check for duplicate code (excluding current record)
            if EmailTemplateType.objects.filter(code=code).exclude(id=template_type_id).exists():
                messages.error(request, 'A template type with this code already exists!')
                return redirect('manage_template_types')
            
            # Update template type
            template_type.name = name
            template_type.code = code
            template_type.description = description
            template_type.is_active = is_active
            template_type.save()
            
            messages.success(
                request, 
                f'Template type "{template_type.name}" has been updated successfully!'
            )
            
        except Exception as e:
            messages.error(request, f'An error occurred while updating: {str(e)}')
    
    return redirect('manage_template_types')


@login_required
@staff_member_required
def delete_template_type(request, template_type_id):
    """Delete template type and all associated templates"""
    template_type = get_object_or_404(EmailTemplateType, id=template_type_id)
    
    if request.method == 'POST':
        try:
            template_name = template_type.name
            templates_count = template_type.templates.count()
            
            # Delete the template type (this will also delete associated templates due to CASCADE)
            template_type.delete()
            
            if templates_count > 0:
                messages.success(
                    request, 
                    f'Template type "{template_name}" and {templates_count} associated template(s) have been deleted successfully!'
                )
            else:
                messages.success(
                    request, 
                    f'Template type "{template_name}" has been deleted successfully!'
                )
            
        except Exception as e:
            messages.error(request, f'An error occurred while deleting: {str(e)}')
    
    return redirect('manage_template_types')


@login_required
@staff_member_required
def toggle_template_type_status(request, template_type_id):
    """Toggle template type active/inactive status"""
    template_type = get_object_or_404(EmailTemplateType, id=template_type_id)
    
    if request.method == 'POST':
        try:
            template_type.is_active = not template_type.is_active
            template_type.save()
            
            status = "activated" if template_type.is_active else "deactivated"
            messages.success(
                request, 
                f'Template type "{template_type.name}" has been {status} successfully!'
            )
            
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
    
    return redirect('manage_template_types')


@login_required
def get_templates_by_type(request, template_type_id):
    """Get templates for a specific template type (AJAX endpoint)"""
    try:
        template_type = get_object_or_404(EmailTemplateType, id=template_type_id)
        templates = template_type.templates.all().order_by('-created_at')
        
        templates_data = []
        for template in templates:
            templates_data.append({
                'id': template.id,
                'name': template.name,
                'subject': template.subject,
                'is_active': template.is_active,
                'created_at': template.created_at.strftime('%b %d, %Y'),
            })
        
        return JsonResponse({
            'success': True,
            'templates': templates_data,
            'template_type': {
                'id': template_type.id,
                'name': template_type.name,
                'code': template_type.code,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


# Update your existing create_email_template view
@login_required 
@staff_member_required
def create_email_template(request):
    """Create new email template with dynamic template types"""
    
    # Get active template types
    template_types = EmailTemplateType.objects.filter(is_active=True).values_list('id', 'name')
    
    # Check if template_type is passed in URL
    selected_template_type = request.GET.get('template_type')
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            template_type_id = request.POST.get('template_type')
            subject = request.POST.get('subject', '').strip()
            email_body = request.POST.get('email_body', '').strip()
            is_active = request.POST.get('is_active') == 'on'
            
            # Validation
            if not all([name, template_type_id, subject, email_body]):
                messages.error(request, 'All fields are required!')
                return render(request, 'email_management/create_template.html', {
                    'template_types': template_types,
                    'selected_template_type': selected_template_type,
                })
            
            # Get template type
            template_type = get_object_or_404(EmailTemplateType, id=template_type_id)
            
            # Create email template
            email_template = EmailTemplate.objects.create(
                name=name,
                template_type=template_type,
                subject=subject,
                email_body=email_body,
                is_active=is_active,
                created_by=request.user
            )
            
            messages.success(
                request, 
                f'Email template "{email_template.name}" has been created successfully!'
            )
            
            return redirect('manage_email_templates')
            
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
    
    context = {
        'template_types': template_types,
        'selected_template_type': selected_template_type,
    }
    
    return render(request, 'email_management/create_template.html', context)



# --------------------------- password reset ---------------------------------
# --------------------------- password reset ---------------------------------
# --------------------------- password reset ---------------------------------
# --------------------------- password reset ---------------------------------
# --------------------------- password reset ---------------------------------

from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_http_methods
from .models import CustomUser, PasswordResetOTP, EmailLog, EmailTemplate, EmailTemplateType
import logging

logger = logging.getLogger(__name__)


@csrf_protect
@require_http_methods(["GET", "POST"])
def password_reset_request(request):
    """View to request password reset OTP"""
    error = None
    success = None
    
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        
        if not email:
            error = "Email address is required."
        else:
            try:
                user = CustomUser.objects.get(email=email)
                
                # Generate OTP
                otp_instance = PasswordResetOTP.generate_otp(user)
                
                # Send OTP email
                if send_otp_email(user, otp_instance.otp):
                    success = f"OTP has been sent to {email}. Please check your inbox."
                    # Store email in session for next step
                    request.session['reset_email'] = email
                else:
                    error = "Failed to send OTP. Please try again."
                    
            except CustomUser.DoesNotExist:
                error = "User with this email does not exist."
    
    return render(request, 'password_reset/request.html', {
        'error': error,
        'success': success
    })


@csrf_protect
@require_http_methods(["GET", "POST"])
def password_reset_verify(request):
    """View to verify OTP"""
    error = None
    success = None
    email = request.session.get('reset_email', '')
    
    if not email:
        messages.error(request, "Session expired. Please request OTP again.")
        return redirect('password_reset_request')
    
    if request.method == 'POST':
        otp = request.POST.get('otp', '').strip()
        
        if not otp:
            error = "OTP is required."
        elif len(otp) != 6 or not otp.isdigit():
            error = "OTP must be 6 digits."
        else:
            try:
                user = CustomUser.objects.get(email=email)
                otp_instance = PasswordResetOTP.objects.get(
                    user=user, 
                    otp=otp, 
                    is_used=False
                )
                
                if otp_instance.is_valid():
                    success = "OTP verified successfully! You can now reset your password."
                    # Store OTP in session for final step
                    request.session['verified_otp'] = otp
                else:
                    error = "OTP has expired. Please request a new one."
                    
            except (CustomUser.DoesNotExist, PasswordResetOTP.DoesNotExist):
                error = "Invalid OTP. Please check and try again."
    
    return render(request, 'password_reset/verify.html', {
        'error': error,
        'success': success,
        'email': email
    })


@csrf_protect
@require_http_methods(["GET", "POST"])
def password_reset_confirm(request):
    """View to confirm new password"""
    error = None
    success = None
    email = request.session.get('reset_email', '')
    verified_otp = request.session.get('verified_otp', '')
    
    if not email or not verified_otp:
        messages.error(request, "Session expired. Please start the reset process again.")
        return redirect('password_reset_request')
    
    if request.method == 'POST':
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')
        
        if not new_password or not confirm_password:
            error = "Both password fields are required."
        elif len(new_password) < 8:
            error = "Password must be at least 8 characters long."
        elif new_password != confirm_password:
            error = "Passwords do not match."
        else:
            try:
                user = CustomUser.objects.get(email=email)
                otp_instance = PasswordResetOTP.objects.get(
                    user=user, 
                    otp=verified_otp, 
                    is_used=False
                )
                
                if otp_instance.is_valid():
                    # Set new password
                    user.set_password(new_password)
                    user.save()
                    
                    # Mark OTP as used
                    otp_instance.is_used = True
                    otp_instance.save()
                    
                    # Send confirmation email
                    send_confirmation_email(user)
                    
                    # Clear session data
                    request.session.pop('reset_email', None)
                    request.session.pop('verified_otp', None)
                    
                    messages.success(request, "Password reset successfully! You can now login with your new password.")
                    return redirect('login')  # Redirect to login page
                else:
                    error = "OTP has expired. Please start the process again."
                    
            except (CustomUser.DoesNotExist, PasswordResetOTP.DoesNotExist):
                error = "Invalid session. Please start the process again."
    
    return render(request, 'password_reset/confirm.html', {
        'error': error,
        'success': success,
        'email': email
    })


def send_otp_email(user, otp):
    """Send OTP via email with proper display name"""
    try:
        # Try to get password reset template
        try:
            template_type = EmailTemplateType.objects.get(code='password_reset')
            email_template = EmailTemplate.objects.filter(
                template_type=template_type, 
                is_active=True
            ).first()
        except (EmailTemplateType.DoesNotExist, EmailTemplate.DoesNotExist):
            email_template = None
        
        if email_template:
            # Use dynamic template
            subject = email_template.subject.format(
                username=user.username,
                first_name=user.first_name or user.username,
                last_name=user.last_name or '',
                email=user.email,
                otp=otp
            )
            
            message = email_template.email_body.format(
                username=user.username,
                first_name=user.first_name or user.username,
                last_name=user.last_name or '',
                email=user.email,
                otp=otp
            )
        else:
            # Fallback to default template
            subject = f'Password Reset OTP - {otp}'
            message = f"""
Dear {user.first_name or user.username},

Your password reset OTP is: {otp}

This OTP is valid for 5 minutes only.

If you didn't request this password reset, please ignore this email.

Best regards,
LMS Support Team
"""
        
        # Create EmailMessage with proper display name
        from_email = f"LMS System <{settings.EMAIL_HOST_USER}>"
        
        email_msg = EmailMessage(
            subject=subject,
            body=message,
            from_email=from_email,
            to=[user.email],
        )
        
        # Send email
        email_msg.send(fail_silently=False)
        
        # Log email
        EmailLog.objects.create(
            recipient_email=user.email,
            recipient_user=user,
            template_used=email_template,
            template_type_used=email_template.template_type if email_template else None,
            subject=subject,
            email_body=message,
            is_sent_successfully=True,
            sent_by=None  # System generated
        )
        
        return True
        
    except Exception as e:
        logger.error(f"Failed to send OTP email to {user.email}: {str(e)}")
        
        # Log failed email
        EmailLog.objects.create(
            recipient_email=user.email,
            recipient_user=user,
            template_used=email_template if 'email_template' in locals() else None,
            template_type_used=email_template.template_type if 'email_template' in locals() and email_template else None,
            subject=subject if 'subject' in locals() else 'Password Reset OTP',
            email_body=message if 'message' in locals() else '',
            is_sent_successfully=False,
            error_message=str(e),
            sent_by=None
        )
        
        return False


def send_confirmation_email(user):
    """Send password reset confirmation email with proper display name"""
    try:
        # Try to get password reset confirmation template
        try:
            template_type = EmailTemplateType.objects.get(code='password_reset_confirmation')
            email_template = EmailTemplate.objects.filter(
                template_type=template_type, 
                is_active=True
            ).first()
        except (EmailTemplateType.DoesNotExist, EmailTemplate.DoesNotExist):
            email_template = None
        
        if email_template:
            # Use dynamic template
            subject = email_template.subject.format(
                username=user.username,
                first_name=user.first_name or user.username,
                last_name=user.last_name or '',
                email=user.email
            )
            
            message = email_template.email_body.format(
                username=user.username,
                first_name=user.first_name or user.username,
                last_name=user.last_name or '',
                email=user.email
            )
        else:
            # Fallback to default template
            subject = 'Password Reset Successful'
            message = f"""
Dear {user.first_name or user.username},

Your password has been successfully reset.

If you didn't make this change, please contact our support team immediately.

Best regards,
LMS Support Team
"""
        
        # Create EmailMessage with proper display name
        from_email = f"LMS System <{settings.EMAIL_HOST_USER}>"
        
        email_msg = EmailMessage(
            subject=subject,
            body=message,
            from_email=from_email,
            to=[user.email],
        )
        
        # Send email
        email_msg.send(fail_silently=True)  # Don't fail if confirmation email fails
        
        # Log email
        EmailLog.objects.create(
            recipient_email=user.email,
            recipient_user=user,
            template_used=email_template,
            template_type_used=email_template.template_type if email_template else None,
            subject=subject,
            email_body=message,
            is_sent_successfully=True,
            sent_by=None
        )
        
    except Exception as e:
        logger.error(f"Failed to send confirmation email to {user.email}: {str(e)}")
        # Don't raise error as password reset was successful



# userss/views.py - Add these instructor views at the end of your existing views.py

# Add these imports at the top with your existing imports
from django.db.models import Count, Q, Sum, Avg
from django.utils import timezone
from datetime import datetime, timedelta

# Add these after your existing views

# ==================== INSTRUCTOR DASHBOARD VIEWS ====================




# permssion views for instructor
# userss/admin_views.py (ya views.py mein add karo)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import JsonResponse
from .models import CustomUser, InstructorPermission, InstructorPermissionAssignment, InstructorProfile
from .decorators import superadmin_required

@superadmin_required
def manage_instructor_permissions(request):
    """
    Admin view to manage all instructor permissions
    """
    # Get all instructors
    instructors = CustomUser.objects.filter(role='instructor').select_related('instructor_profile')
    
    # Get all available permissions
    all_permissions = InstructorPermission.objects.filter(is_active=True).order_by('category', 'name')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        instructors = instructors.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Filter by permission
    permission_filter = request.GET.get('permission', '')
    if permission_filter:
        instructors = instructors.filter(
            instructor_profile__permissions__permission__code=permission_filter,
            instructor_profile__permissions__is_active=True
        ).distinct()
    
    context = {
        'instructors': instructors,
        'all_permissions': all_permissions,
        'search_query': search_query,
        'permission_filter': permission_filter,
    }
    
    return render(request, 'admin/manage_instructor_permissions.html', context)



# views.py mein instructor_permission_detail function update karo:

@superadmin_required
def instructor_permission_detail(request, instructor_id):
    """
    Detailed view for managing single instructor's permissions
    """
    instructor = get_object_or_404(CustomUser, id=instructor_id, role='instructor')
    
    # Get or create instructor profile
    instructor_profile, created = InstructorProfile.objects.get_or_create(
        user=instructor,
        defaults={'created_by': request.user}
    )
    
    # Get all available permissions
    all_permissions = InstructorPermission.objects.filter(is_active=True).order_by('category', 'name')
    
    # Get instructor's current permissions
    current_permissions = InstructorPermissionAssignment.objects.filter(
        instructor=instructor_profile,
        is_active=True
    ).select_related('permission')
    
    # Create a dictionary for easy lookup WITH ASSIGNMENT DETAILS
    current_permission_data = {}
    for assignment in current_permissions:
        current_permission_data[assignment.permission.code] = {
            'assignment': assignment,
            'assigned_date': assignment.assigned_at
        }
    
    context = {
        'instructor': instructor,
        'instructor_profile': instructor_profile,
        'all_permissions': all_permissions,
        'current_permission_codes': [p.permission.code for p in current_permissions],
        'current_permission_data': current_permission_data,
    }
    
    return render(request, 'admin/instructor_permission_detail.html', context)


@superadmin_required
def update_instructor_permissions(request, instructor_id):
    """
    Update instructor permissions via POST request
    """
    if request.method != 'POST':
        return redirect('manage_instructor_permissions')
    
    instructor = get_object_or_404(CustomUser, id=instructor_id, role='instructor')
    instructor_profile, created = InstructorProfile.objects.get_or_create(
        user=instructor,
        defaults={'created_by': request.user}
    )
    
    # Get selected permissions from form
    selected_permissions = request.POST.getlist('permissions')
    
    # Get all available permissions
    all_permissions = InstructorPermission.objects.filter(is_active=True)
    
    # Deactivate all current permissions
    InstructorPermissionAssignment.objects.filter(
        instructor=instructor_profile
    ).update(is_active=False)
    
    # Add new permissions
    assigned_count = 0
    for permission in all_permissions:
        if permission.code in selected_permissions:
            assignment, created = InstructorPermissionAssignment.objects.get_or_create(
                instructor=instructor_profile,
                permission=permission,
                defaults={
                    'assigned_by': request.user,
                    'is_active': True
                }
            )
            if not created:
                assignment.is_active = True
                assignment.assigned_by = request.user
                assignment.save()
            assigned_count += 1
    
    messages.success(
        request, 
        f'Successfully updated permissions for {instructor.get_full_name()}. '
        f'{assigned_count} permissions assigned.'
    )
    
    return redirect('instructor_permission_detail', instructor_id=instructor_id)

@superadmin_required
def bulk_permission_update(request):
    """
    Bulk update permissions for multiple instructors
    """
    if request.method != 'POST':
        return redirect('manage_instructor_permissions')
    
    instructor_ids = request.POST.getlist('instructor_ids')
    permission_codes = request.POST.getlist('bulk_permissions')
    action = request.POST.get('bulk_action')  # 'add' or 'remove'
    
    if not instructor_ids or not permission_codes:
        messages.error(request, 'Please select instructors and permissions.')
        return redirect('manage_instructor_permissions')
    
    # Get instructors and permissions
    instructors = CustomUser.objects.filter(id__in=instructor_ids, role='instructor')
    permissions = InstructorPermission.objects.filter(code__in=permission_codes, is_active=True)
    
    updated_count = 0
    
    for instructor in instructors:
        instructor_profile, created = InstructorProfile.objects.get_or_create(
            user=instructor,
            defaults={'created_by': request.user}
        )
        
        for permission in permissions:
            if action == 'add':
                assignment, created = InstructorPermissionAssignment.objects.get_or_create(
                    instructor=instructor_profile,
                    permission=permission,
                    defaults={
                        'assigned_by': request.user,
                        'is_active': True
                    }
                )
                if not created and not assignment.is_active:
                    assignment.is_active = True
                    assignment.assigned_by = request.user
                    assignment.save()
                    
            elif action == 'remove':
                InstructorPermissionAssignment.objects.filter(
                    instructor=instructor_profile,
                    permission=permission
                ).update(is_active=False)
        
        updated_count += 1
    
    action_text = 'assigned' if action == 'add' else 'removed'
    messages.success(
        request,
        f'Successfully {action_text} permissions for {updated_count} instructors.'
    )
    
    return redirect('manage_instructor_permissions')

@superadmin_required
def create_instructor_permission(request):
    """
    Create new instructor permission
    """
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        description = request.POST.get('description', '').strip()
        category = request.POST.get('category', '').strip()
        
        if not all([name, code, description, category]):
            messages.error(request, 'All fields are required.')
        elif InstructorPermission.objects.filter(code=code).exists():
            messages.error(request, 'Permission code already exists.')
        else:
            InstructorPermission.objects.create(
                name=name,
                code=code,
                description=description,
                category=category,
                created_by=request.user
            )
            messages.success(request, f'Permission "{name}" created successfully.')
            return redirect('manage_instructor_permissions')
    
    context = {
        'categories': InstructorPermission.objects.values_list('category', flat=True).distinct()
    }
    return render(request, 'admin/create_instructor_permission.html', context)







# views.py (instructor dashboard views)
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Count, Q
from .decorators import instructor_required, instructor_permission_required
from userss.models import CustomUser
from courses.models import Course, Enrollment, Batch

@instructor_required
def instructor_dashboard(request):
    """
    Main instructor dashboard
    """
    instructor = request.user
    
    # Get instructor's courses
    instructor_courses = Course.objects.filter(
        Q(instructor=instructor) | Q(co_instructors=instructor)
    ).distinct()
    
    # Get basic stats
    total_courses = instructor_courses.count()
    total_students = Enrollment.objects.filter(
        course__in=instructor_courses,
        is_active=True
    ).count()
    
    # Get recent enrollments
    recent_enrollments = Enrollment.objects.filter(
        course__in=instructor_courses,
        is_active=True
    ).select_related('student', 'course').order_by('-enrolled_at')[:5]
    
    # Get instructor permissions
    permissions = instructor.get_instructor_permissions()
    
    # ⭐ EXACT DATABASE CODE MAPPING ⭐
    permission_mapping = {
        # Basic
        'course_management': 'has_course_permission',
        'batch_management': 'has_batch_permission',
        'content_management': 'has_content_permission',
        
        # Exam
        'exam_dashboard': 'has_exam_permission',
        'create_exam': 'has_create_exam_permission',
        'my_exams': 'has_my_exams_permission',
        'assign_exam': 'has_assign_exam_permission',
        
        # Attendance
        'attendance_dashboard': 'has_attendance_dashboard_permission',
        'createsession': 'has_createsession_permission',  # ⭐ NO UNDERSCORE
        'all_sessions': 'has_all_sessions_permission',    # ⭐ PLURAL
        'attendance_pending': 'has_attendance_pending_permission',  # ⭐ SHORTER
        'attendance_reports': 'has_attendance_reports_permission',
        
        # Other
        'student_management': 'has_student_permission',
        'email_marketing': 'has_email_permission',
        'profile_setting': 'has_profile_setting_permission',
    }
    
    # Generate flags
    permission_flags = {}
    for perm_code, template_var in permission_mapping.items():
        permission_flags[template_var] = instructor.has_instructor_permission(perm_code)
    
    context = {
        'total_courses': total_courses,
        'total_students': total_students,
        'recent_enrollments': recent_enrollments,
        'instructor_courses': instructor_courses[:5],
        'permissions': permissions,
        **permission_flags,
    }
    
    return render(request, 'instructor/instructor_dashboard.html', context)

@instructor_permission_required('course_management')
def instructor_courses(request):
    """
    View for managing instructor's courses
    """
    instructor = request.user
    
    courses = Course.objects.filter(
        Q(instructor=instructor) | Q(co_instructors=instructor)
    ).distinct().select_related('category').prefetch_related('enrollments')
    
    context = {
        'courses': courses,
    }
    
    return render(request, 'instructor/courses.html', context)

@instructor_permission_required('content_management')
def content_management(request):
    """
    View for managing course content
    """
    instructor = request.user
    
    courses = Course.objects.filter(
        Q(instructor=instructor) | Q(co_instructors=instructor)
    ).distinct()
    
    context = {
        'courses': courses,
    }
    
    return render(request, 'instructor/content_management.html', context)

@instructor_permission_required('batch_management')
def batch_management(request):
    """
    View for managing batches
    """
    instructor = request.user
    
    # Get instructor's courses
    instructor_courses = Course.objects.filter(
        Q(instructor=instructor) | Q(co_instructors=instructor)
    ).distinct()
    
    # Get batches for instructor's courses
    batches = Batch.objects.filter(
        course__in=instructor_courses
    ).select_related('course').order_by('-created_at')
    
    context = {
        'batches': batches,
        'instructor_courses': instructor_courses,
    }
    
    return render(request, 'instructor/batch_management.html', context)

@instructor_permission_required('student_management')
def student_management(request):
    """
    View for managing students
    """
    instructor = request.user
    
    # Get instructor's courses
    instructor_courses = Course.objects.filter(
        Q(instructor=instructor) | Q(co_instructors=instructor)
    ).distinct()
    
    # Get students enrolled in instructor's courses
    enrollments = Enrollment.objects.filter(
        course__in=instructor_courses,
        is_active=True
    ).select_related('student', 'course').order_by('-enrolled_at')
    
    context = {
        'enrollments': enrollments,
        'instructor_courses': instructor_courses,
    }
    
    return render(request, 'instructor/student_management.html', context)



@instructor_permission_required('email_marketing')
def email_marketing(request):
    """
    View for email marketing
    """
    instructor = request.user
    
    # Get instructor's courses for email targeting
    instructor_courses = Course.objects.filter(
        Q(instructor=instructor) | Q(co_instructors=instructor)
    ).distinct()
    
    context = {
        'instructor_courses': instructor_courses,
    }
    
    return render(request, 'instructor/email_marketing.html', context)

@instructor_permission_required('analytics_view')
def course_analytics(request):
    """
    View for course analytics
    """
    instructor = request.user
    
    # Get instructor's courses with analytics data
    courses_data = []
    instructor_courses = Course.objects.filter(
        Q(instructor=instructor) | Q(co_instructors=instructor)
    ).distinct()
    
    for course in instructor_courses:
        enrollments = course.enrollments.filter(is_active=True)
        courses_data.append({
            'course': course,
            'total_enrollments': enrollments.count(),
            'completed_enrollments': enrollments.filter(status='completed').count(),
            'active_enrollments': enrollments.filter(status='enrolled').count(),
        })
    
    context = {
        'courses_data': courses_data,
    }
    
    return render(request, 'instructor/analytics.html', context)

@instructor_required
def check_permission(request):
    """
    AJAX endpoint to check if instructor has specific permission
    """
    permission_code = request.GET.get('permission')
    if not permission_code:
        return JsonResponse({'error': 'Permission code required'}, status=400)
    
    has_permission = request.user.has_instructor_permission(permission_code)
    
    return JsonResponse({
        'has_permission': has_permission,
        'permission_code': permission_code
    })
from courses.models import *



# userss/views.py (ya jahan aapka instructor views hain)

@login_required
def instructor_content_management(request):
    """Content Management - Show only assigned courses to instructor"""
    if request.user.role != 'instructor':
        return redirect('user_login')
    
    # Get only courses assigned to this instructor
    courses = Course.objects.filter(
        instructor=request.user,
        is_active=True
    ).prefetch_related('modules', 'batches', 'enrollments', 'category')
    
    # Calculate stats for instructor's assigned courses only
    total_batches = Batch.objects.filter(course__instructor=request.user).count()
    total_students = BatchEnrollment.objects.filter(
        batch__course__instructor=request.user,
        is_active=True
    ).count()
    
    # Categories for filter (optional)
    categories = CourseCategory.objects.filter(is_active=True)
    
    context = {
        'courses': courses,  # Yeh zaroori hai!
        'total_courses': courses.count(),
        'published_courses': courses.filter(status='published').count(),
        'total_batches': total_batches,
        'total_students': total_students,
        'categories': categories,
    }
    
    return render(request, 'instructor/content_management.html', context)

@login_required
def instructor_batch_management(request):
    """Instructor: Main batch management dashboard"""
    if request.user.role != 'instructor':
        return redirect('user_login')
    
    # ✅ YEH CHANGE - Batch ke instructor ke unique courses
    all_batches = Batch.objects.filter(
        instructor=request.user
    ).select_related('course', 'instructor').order_by('-created_at')
    
    # ✅ Unique courses from batches (jinke batches instructor ne banaye hain)
    instructor_courses = Course.objects.filter(
        batches__instructor=request.user
    ).distinct()
    
    # Recent batches (latest 5 for activity section)
    recent_batches = all_batches[:5]
    
    # Calculate stats
    stats = {
        'total_batches': all_batches.count(),
        'active_batches': all_batches.filter(status='active').count(),
        'total_students': BatchEnrollment.objects.filter(
            batch__instructor=request.user,
            is_active=True
        ).count(),
        'total_courses': instructor_courses.count(),
    }
    
    # Add enrolled count to each batch
    for batch in recent_batches:
        batch.enrolled_count = batch.get_enrolled_count()
    
    context = {
        'batches': recent_batches,
        'stats': stats,
        'user': request.user,
        'instructor_courses': instructor_courses,  # ✅ YEH ADD KARO
    }
    
    return render(request, 'instructor/batch_management.html', context)




from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, Avg, Prefetch
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import json
from courses.models import Course, Batch, Enrollment, BatchEnrollment
from .models import CustomUser

@login_required
def instructor_student_management(request):
    """Enhanced student management with filtering, search, and pagination"""
    
    # Check if user is instructor
    if request.user.role != 'instructor':
        return redirect('dashboard')
    
    # Get instructor's courses
    instructor_courses = Course.objects.filter(
        instructor=request.user,
        is_active=True
    ).prefetch_related('batches')
    
    # Get all enrollments for instructor's courses
    enrollments = Enrollment.objects.filter(
        course__instructor=request.user
    ).select_related(
        'student', 'course'
    ).prefetch_related(
        'student__batch_enrollments__batch'
    ).order_by('-enrolled_at')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        enrollments = enrollments.filter(
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__email__icontains=search_query) |
            Q(student__username__icontains=search_query) |
            Q(course__title__icontains=search_query) |
            Q(course__course_code__icontains=search_query)
        )
    
    # Course filter
    course_filter = request.GET.get('course')
    if course_filter:
        enrollments = enrollments.filter(course_id=course_filter)
    
    # Batch filter
    batch_filter = request.GET.get('batch')
    available_batches = []
    if course_filter:
        available_batches = Batch.objects.filter(
            course_id=course_filter,
            instructor=request.user
        )
        if batch_filter:
            # Filter enrollments by batch
            batch_enrollment_students = BatchEnrollment.objects.filter(
                batch_id=batch_filter
            ).values_list('student_id', flat=True)
            enrollments = enrollments.filter(student_id__in=batch_enrollment_students)
    else:
        available_batches = Batch.objects.filter(
            instructor=request.user
        )
    
    # Status filter
    status_filter = request.GET.get('status')
    if status_filter:
        enrollments = enrollments.filter(status=status_filter)
    
    # Sorting
    sort_by = request.GET.get('sort_by', 'recent')
    if sort_by == 'name':
        enrollments = enrollments.order_by('student__first_name', 'student__last_name')
    elif sort_by == 'progress':
        enrollments = enrollments.order_by('-progress_percentage')
    elif sort_by == 'course':
        enrollments = enrollments.order_by('course__course_code')
    else:  # recent
        enrollments = enrollments.order_by('-enrolled_at')
    
    # Add batch enrollment info to each enrollment
    for enrollment in enrollments:
        # Get the batch enrollment for this student in this course
        batch_enrollment = BatchEnrollment.objects.filter(
            student=enrollment.student,
            batch__course=enrollment.course
        ).select_related('batch').first()
        enrollment.batch_enrollment = batch_enrollment
    
    # Statistics
    total_students = enrollments.values('student').distinct().count()
    active_students = enrollments.filter(is_active=True).values('student').distinct().count()
    total_courses = instructor_courses.count()
    total_batches = Batch.objects.filter(instructor=request.user, is_active=True).count()
    
    # Pagination
    paginator = Paginator(enrollments, 20)  # 20 students per page
    page_number = request.GET.get('page')
    page_enrollments = paginator.get_page(page_number)
    
    context = {
        'enrollments': page_enrollments,
        'my_courses': instructor_courses,
        'available_batches': available_batches,
        'total_students': total_students,
        'active_students': active_students,
        'total_courses': total_courses,
        'total_batches': total_batches,
        'search_query': search_query,
    }
    
    return render(request, 'instructor/student_management.html', context)


# API endpoints for AJAX functionality

@login_required
@require_http_methods(["GET"])
def get_course_batches_api(request, course_id):
    """API to get batches for a specific course"""
    if request.user.role != 'instructor':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        course = Course.objects.get(id=course_id, instructor=request.user)
        batches = course.batches.filter(is_active=True).values('id', 'name', 'code')
        return JsonResponse({'batches': list(batches)})
    except Course.DoesNotExist:
        return JsonResponse({'error': 'Course not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def toggle_enrollment_status(request, enrollment_id):
    """Toggle enrollment active status"""
    if request.user.role != 'instructor':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        enrollment = Enrollment.objects.get(
            id=enrollment_id,
            course__instructor=request.user
        )
        enrollment.is_active = not enrollment.is_active
        enrollment.save()
        
        return JsonResponse({
            'success': True,
            'new_status': enrollment.is_active,
            'message': f'Student {"activated" if enrollment.is_active else "deactivated"} successfully'
        })
    except Enrollment.DoesNotExist:
        return JsonResponse({'error': 'Enrollment not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def change_enrollment_status(request, enrollment_id):
    """Change enrollment status (completed, suspended, etc.)"""
    if request.user.role != 'instructor':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if new_status not in ['enrolled', 'completed', 'dropped', 'suspended']:
            return JsonResponse({'error': 'Invalid status'}, status=400)
        
        enrollment = Enrollment.objects.get(
            id=enrollment_id,
            course__instructor=request.user
        )
        enrollment.status = new_status
        
        # Set completion date if marked as completed
        if new_status == 'completed' and not enrollment.completed_at:
            from django.utils import timezone
            enrollment.completed_at = timezone.now()
        
        enrollment.save()
        
        return JsonResponse({
            'success': True,
            'new_status': new_status,
            'message': f'Student status changed to {new_status} successfully'
        })
    except Enrollment.DoesNotExist:
        return JsonResponse({'error': 'Enrollment not found'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)


@login_required
@require_http_methods(["POST"])
def bulk_change_enrollment_status(request):
    """Bulk change status for multiple enrollments"""
    if request.user.role != 'instructor':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        data = json.loads(request.body)
        enrollment_ids = data.get('enrollment_ids', [])
        new_status = data.get('status')
        
        if not enrollment_ids:
            return JsonResponse({'error': 'No enrollments selected'}, status=400)
        
        if new_status not in ['enrolled', 'completed', 'dropped', 'suspended']:
            return JsonResponse({'error': 'Invalid status'}, status=400)
        
        # Update enrollments
        enrollments = Enrollment.objects.filter(
            id__in=enrollment_ids,
            course__instructor=request.user
        )
        
        updated_count = 0
        for enrollment in enrollments:
            enrollment.status = new_status
            if new_status == 'completed' and not enrollment.completed_at:
                from django.utils import timezone
                enrollment.completed_at = timezone.now()
            enrollment.save()
            updated_count += 1
        
        return JsonResponse({
            'success': True,
            'updated_count': updated_count,
            'message': f'{updated_count} students updated successfully'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)


@login_required
@require_http_methods(["POST"])
def remove_student_enrollment(request, enrollment_id):
    """Remove student from course (soft delete)"""
    if request.user.role != 'instructor':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        enrollment = Enrollment.objects.get(
            id=enrollment_id,
            course__instructor=request.user
        )
        
        # Soft delete - mark as inactive and dropped
        enrollment.is_active = False
        enrollment.status = 'dropped'
        enrollment.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Student removed successfully'
        })
    except Enrollment.DoesNotExist:
        return JsonResponse({'error': 'Enrollment not found'}, status=404)





# Add these views to your existing views.py for instructor email functionality

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.http import JsonResponse
from django.db.models import Q
from datetime import date
from .models import CustomUser, EmailTemplate, EmailTemplateType, EmailLog, DailyEmailSummary, EmailLimitSet
from courses.models import Batch, BatchEnrollment

@login_required
def instructor_email_management(request):
    """Main email management page for instructors"""
    if request.user.role != 'instructor':
        messages.error(request, 'Access denied. Instructors only.')
        return redirect('user_login')
    
    # Get instructor's batches
    instructor_batches = Batch.objects.filter(
        instructor=request.user,
        is_active=True
    ).select_related('course')
    
    # Get active email templates
    email_templates = EmailTemplate.objects.filter(
        is_active=True
    ).select_related('template_type')
    
    # Get current email limit info
    email_limit_setting = EmailLimitSet.objects.filter(is_active=True).first()
    today_summary = DailyEmailSummary.objects.filter(date=date.today()).first()
    
    # Calculate remaining emails for today
    if email_limit_setting and today_summary:
        remaining_emails = email_limit_setting.email_limit_per_day - today_summary.total_emails_sent
        remaining_emails = max(0, remaining_emails)
    else:
        remaining_emails = email_limit_setting.email_limit_per_day if email_limit_setting else 50
    
    # Get recent email logs for this instructor
    recent_emails = EmailLog.objects.filter(
        sent_by=request.user
    ).order_by('-sent_time')[:5]
    
    context = {
        'instructor_batches': instructor_batches,
        'email_templates': email_templates,
        'email_limit_setting': email_limit_setting,
        'today_summary': today_summary,
        'remaining_emails': remaining_emails,
        'recent_emails': recent_emails,
    }
    
    return render(request, 'instructor/email_management.html', context)


@login_required
def instructor_send_batch_email(request):
    """Send email to batch students"""
    if request.user.role != 'instructor':
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    if request.method == 'POST':
        template_id = request.POST.get('template_id')
        batch_id = request.POST.get('batch_id')
        selected_students = request.POST.getlist('selected_students')
        custom_subject = request.POST.get('custom_subject', '').strip()
        custom_message = request.POST.get('custom_message', '').strip()
        
        # Validation
        if not template_id and (not custom_subject or not custom_message):
            messages.error(request, 'Please select a template or provide custom subject and message.')
            return redirect('instructor_email_management')
        
        if not batch_id:
            messages.error(request, 'Please select a batch.')
            return redirect('instructor_email_management')
        
        # Get batch and verify instructor owns it
        try:
            batch = Batch.objects.get(id=batch_id, instructor=request.user, is_active=True)
        except Batch.DoesNotExist:
            messages.error(request, 'Invalid batch selection.')
            return redirect('instructor_email_management')
        
        # Check daily email limit
        if check_daily_email_limit():
            messages.error(request, 'Daily email limit reached! Cannot send emails.')
            return redirect('instructor_email_management')
        
        # Get email template if selected
        template = None
        if template_id:
            try:
                template = EmailTemplate.objects.get(id=template_id, is_active=True)
            except EmailTemplate.DoesNotExist:
                messages.error(request, 'Invalid template selection.')
                return redirect('instructor_email_management')
        
        # Get recipients based on selection
        if selected_students:
            # Specific students selected
            recipients = CustomUser.objects.filter(
                id__in=selected_students,
                role='student',
                batch_enrollments__batch=batch,
                batch_enrollments__is_active=True,
                email__isnull=False
            ).exclude(email='').distinct()
        else:
            # All batch students
            recipients = CustomUser.objects.filter(
                role='student',
                batch_enrollments__batch=batch,
                batch_enrollments__is_active=True,
                email__isnull=False
            ).exclude(email='').distinct()
        
        if not recipients.exists():
            messages.error(request, 'No valid recipients found in the selected batch.')
            return redirect('instructor_email_management')
        
        # Send emails
        successful_count = 0
        failed_count = 0
        failed_emails = []
        
        for student in recipients:
            # Check individual email limit before each send
            if check_daily_email_limit():
                messages.warning(request, f'Daily email limit reached! Only {successful_count} emails were sent.')
                break
            
            try:
                # Use template or custom content
                if template:
                    subject = template.subject
                    message = template.email_body
                else:
                    subject = custom_subject
                    message = custom_message
                
                # Replace template variables
                context_vars = {
                    '{{username}}': student.username or student.email.split('@')[0],
                    '{{email}}': student.email or '',
                    '{{first_name}}': student.first_name or student.username or student.email.split('@')[0],
                    '{{last_name}}': student.last_name or '',
                    '{{batch_name}}': batch.name,
                    '{{course_name}}': batch.course.title,
                    '{{instructor_name}}': request.user.get_full_name() or request.user.username,
                }
                
                # Replace variables in subject and message
                for var, value in context_vars.items():
                    subject = subject.replace(var, str(value))
                    message = message.replace(var, str(value))
                
                # Send email
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[student.email],
                    fail_silently=False
                )
                
                successful_count += 1
                
                # Log successful email
                EmailLog.objects.create(
                    recipient_email=student.email,
                    recipient_user=student,
                    template_used=template,
                    template_type_used=template.template_type if template else None,
                    subject=subject,
                    email_body=message,
                    is_sent_successfully=True,
                    sent_by=request.user
                )
                
            except Exception as e:
                failed_count += 1
                failed_emails.append(f"{student.email}: {str(e)}")
                
                # Log failed email
                EmailLog.objects.create(
                    recipient_email=student.email,
                    recipient_user=student,
                    template_used=template,
                    template_type_used=template.template_type if template else None,
                    subject=custom_subject if not template else template.subject,
                    email_body=custom_message if not template else template.email_body,
                    is_sent_successfully=False,
                    error_message=str(e),
                    sent_by=request.user
                )
        
        # Update daily summary
        try:
            today = date.today()
            email_limit = 50  # Default
            if EmailLimitSet.objects.filter(is_active=True).exists():
                email_limit = EmailLimitSet.objects.filter(is_active=True).first().email_limit_per_day
            
            daily_summary, created = DailyEmailSummary.objects.get_or_create(
                date=today,
                defaults={'daily_limit': email_limit}
            )
            daily_summary.total_emails_sent += (successful_count + failed_count)
            daily_summary.successful_emails += successful_count
            daily_summary.failed_emails += failed_count
            daily_summary.save()
        except Exception as e:
            pass  # Don't break the flow for logging errors
        
        # Show results
        if successful_count > 0:
            messages.success(request, f'Email sent successfully to {successful_count} students in {batch.name}!')
        
        if failed_count > 0:
            messages.warning(request, f'{failed_count} emails failed to send.')
        
        if successful_count == 0 and failed_count == 0:
            messages.error(request, 'No emails were sent!')
        
        return redirect('instructor_email_management')
    
    return redirect('instructor_email_management')


@login_required
def get_batch_students(request, batch_id):
    """AJAX endpoint to get students in a specific batch"""
    if request.user.role != 'instructor':
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Verify instructor owns this batch
        batch = Batch.objects.get(id=batch_id, instructor=request.user, is_active=True)
        
        # Get enrolled students
        students = CustomUser.objects.filter(
            role='student',
            batch_enrollments__batch=batch,
            batch_enrollments__is_active=True,
            email__isnull=False
        ).exclude(email='').distinct().values(
            'id', 'username', 'first_name', 'last_name', 'email'
        )
        
        students_list = []
        for student in students:
            students_list.append({
                'id': student['id'],
                'name': f"{student['first_name']} {student['last_name']}".strip() or student['username'],
                'email': student['email']
            })
        
        return JsonResponse({
            'success': True,
            'students': students_list,
            'batch_name': batch.name,
            'total_count': len(students_list)
        })
        
    except Batch.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Batch not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def get_email_template_preview(request, template_id):
    """AJAX endpoint to preview email template"""
    if request.user.role != 'instructor':
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        template = EmailTemplate.objects.get(id=template_id, is_active=True)
        
        # Sample data for preview
        sample_subject = template.subject.replace('{{username}}', 'John Doe')
        sample_subject = sample_subject.replace('{{first_name}}', 'John')
        sample_subject = sample_subject.replace('{{last_name}}', 'Doe')
        sample_subject = sample_subject.replace('{{batch_name}}', 'Sample Batch')
        sample_subject = sample_subject.replace('{{course_name}}', 'Sample Course')
        sample_subject = sample_subject.replace('{{instructor_name}}', request.user.get_full_name() or request.user.username)
        
        sample_body = template.email_body.replace('{{username}}', 'John Doe')
        sample_body = sample_body.replace('{{first_name}}', 'John')
        sample_body = sample_body.replace('{{last_name}}', 'Doe')
        sample_body = sample_body.replace('{{batch_name}}', 'Sample Batch')
        sample_body = sample_body.replace('{{course_name}}', 'Sample Course')
        sample_body = sample_body.replace('{{instructor_name}}', request.user.get_full_name() or request.user.username)
        
        return JsonResponse({
            'success': True,
            'template': {
                'id': template.id,
                'name': template.name,
                'subject': sample_subject,
                'body': sample_body,
                'original_subject': template.subject,
                'original_body': template.email_body
            }
        })
        
    except EmailTemplate.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Template not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def instructor_email_history(request):
    """View instructor's email sending history"""
    if request.user.role != 'instructor':
        messages.error(request, 'Access denied. Instructors only.')
        return redirect('user_login')
    
    # Get email logs for this instructor
    email_logs = EmailLog.objects.filter(
        sent_by=request.user
    ).select_related('template_used', 'template_type_used', 'recipient_user').order_by('-sent_time')
    
    # Filter by date if provided
    date_filter = request.GET.get('date')
    if date_filter:
        email_logs = email_logs.filter(sent_date=date_filter)
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter == 'success':
        email_logs = email_logs.filter(is_sent_successfully=True)
    elif status_filter == 'failed':
        email_logs = email_logs.filter(is_sent_successfully=False)
    
    # Statistics
    total_sent = email_logs.count()
    successful_sent = email_logs.filter(is_sent_successfully=True).count()
    failed_sent = email_logs.filter(is_sent_successfully=False).count()
    success_rate = round((successful_sent / total_sent * 100), 1) if total_sent > 0 else 0
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(email_logs, 20)
    page = request.GET.get('page')
    email_logs = paginator.get_page(page)
    
    context = {
        'email_logs': email_logs,
        'total_sent': total_sent,
        'successful_sent': successful_sent,
        'failed_sent': failed_sent,
        'success_rate': success_rate,
        'date_filter': date_filter,
        'status_filter': status_filter,
    }
    
    return render(request, 'instructor/email_history.html', context)





def instructor_profile_management(request):
    return render(request,'instructor/profile_management.html')



# courses/views.py - Add this simple view

@login_required
def instructor_course_management(request):
    """Simple view for instructor to see their courses - NO CREATE/EDIT options"""
    if request.user.role != 'instructor':
        return redirect('user_login')
    
    # Get instructor's courses with basic stats
    courses = Course.objects.filter(
        Q(instructor=request.user) | Q(co_instructors=request.user)
    ).distinct().select_related('category').annotate(
        enrollment_count=Count('enrollments', filter=Q(enrollments__is_active=True)),
        module_count=Count('modules', filter=Q(modules__is_active=True))
    ).order_by('-created_at')
    
    # Basic search functionality (optional)
    search_query = request.GET.get('search', '')
    if search_query:
        courses = courses.filter(
            Q(title__icontains=search_query) |
            Q(course_code__icontains=search_query)
        )
    
    # Calculate basic stats
    total_courses = courses.count()
    published_courses = courses.filter(status='published').count()
    draft_courses = courses.filter(status='draft').count()
    total_students = sum([course.enrollment_count for course in courses])
    
    context = {
        'courses': courses,
        'search_query': search_query,
        'total_courses': total_courses,
        'published_courses': published_courses,
        'draft_courses': draft_courses,
        'total_students': total_students,
    }
    
    return render(request, 'instructor/course_management.html', context)


# courses/views.py - Add these views

@login_required
def instructor_view_students(request):
    """Instructor view to see all their students across courses and batches"""
    if request.user.role != 'instructor':
        return redirect('user_login')
    
    # Get instructor's courses
    instructor_courses = Course.objects.filter(
        Q(instructor=request.user) | Q(co_instructors=request.user)
    ).distinct().prefetch_related('enrollments__student', 'batches__enrollments__student')
    
    # Get course enrollments
    course_enrollments = Enrollment.objects.filter(
        course__in=instructor_courses,
        is_active=True
    ).select_related('student', 'course').order_by('-enrolled_at')
    
    # Get batch enrollments  
    batch_enrollments = BatchEnrollment.objects.filter(
        batch__course__in=instructor_courses,
        is_active=True
    ).select_related('student', 'batch', 'batch__course').order_by('-enrolled_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        course_enrollments = course_enrollments.filter(
            Q(student__username__icontains=search_query) |
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__email__icontains=search_query)
        )
        batch_enrollments = batch_enrollments.filter(
            Q(student__username__icontains=search_query) |
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__email__icontains=search_query)
        )
    
    # Calculate stats
    total_course_students = course_enrollments.count()
    total_batch_students = batch_enrollments.count()
    total_courses = instructor_courses.count()
    total_batches = sum([course.batches.count() for course in instructor_courses])
    
    context = {
        'instructor_courses': instructor_courses,
        'course_enrollments': course_enrollments,
        'batch_enrollments': batch_enrollments,
        'search_query': search_query,
        'total_course_students': total_course_students,
        'total_batch_students': total_batch_students,
        'total_courses': total_courses,
        'total_batches': total_batches,
    }
    
    return render(request, 'instructor/view_students.html', context)


@login_required
def instructor_course_detail(request, course_id):
    """Detailed view of specific course with students and batches"""
    if request.user.role != 'instructor':
        return redirect('user_login')
    
    # Get course and check permission
    course = get_object_or_404(Course, id=course_id)
    if course.instructor != request.user and request.user not in course.co_instructors.all():
        messages.error(request, "You don't have permission to view this course!")
        return redirect('instructor_see_all_course')
    
    # Get course data
    modules = course.modules.filter(is_active=True).prefetch_related('lessons')
    enrollments = course.enrollments.filter(is_active=True).select_related('student')
    batches = course.batches.all().prefetch_related('enrollments__student')
    
    # Calculate stats
    total_modules = modules.count()
    total_lessons = sum([module.lessons.filter(is_active=True).count() for module in modules])
    total_students = enrollments.count()
    total_batch_students = sum([batch.enrollments.filter(is_active=True).count() for batch in batches])
    
    context = {
        'course': course,
        'modules': modules,
        'enrollments': enrollments,
        'batches': batches,
        'total_modules': total_modules,
        'total_lessons': total_lessons,
        'total_students': total_students,
        'total_batch_students': total_batch_students,
    }
    
    return render(request, 'instructor/course_detail.html', context)





# exam management
# exam management
# exam management
# exam management
# exam management
# exam management
# exam management

def createExam(request):
    return render(request,"exam/create_exam.html")

def assignExam(request):
    return render(request,"exam/assign_exam.html")

def exam_submission(request):
    return render(request,"exam/exam_submissions.html")

from django.core.mail import send_mail
from django.conf import settings

@login_required
def send_email_to_user(request, user_id):
    if request.user.role != 'superadmin':
        return redirect('user_login')
    
    user = get_object_or_404(CustomUser, id=user_id)
    
    if request.method == 'POST':
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        email_type = request.POST.get('email_type', 'general')
        priority = request.POST.get('priority', 'normal')
        send_copy = request.POST.get('send_copy')
        
        # Check email limit
        today = date.today()
        email_limit = EmailLimitSet.objects.filter(is_active=True).first()
        daily_limit = email_limit.email_limit_per_day if email_limit else 50
        
        today_emails = EmailLog.objects.filter(
            recipient_user=user,
            sent_date=today
        ).count()
        
        if today_emails >= daily_limit:
            messages.warning(request, f'User has reached daily email limit ({daily_limit}). Email not sent.')
            return redirect('user_details', user_id=user_id)
        
        try:
            # Send email
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                fail_silently=False,
            )
            
            # Log email
            EmailLog.objects.create(
                recipient_user=user,
                sent_by=request.user,
                subject=subject,
                email_body=message,
                is_sent_successfully=True,
                sent_date=today,
            )
            
            # Send copy to self if requested
            if send_copy:
                send_mail(
                    subject=f"Copy: {subject}",
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[request.user.email],
                    fail_silently=True,
                )
            
            messages.success(request, f'Email sent successfully to {user.get_full_name()}!')
            return redirect('user_details', user_id=user_id)
            
        except Exception as e:
            # Log failed email
            EmailLog.objects.create(
                recipient_user=user,
                sent_by=request.user,
                subject=subject,
                email_body=message,
                is_sent_successfully=False,
                sent_date=today,
            )
            messages.error(request, f'Failed to send email: {str(e)}')
    
    return render(request, 'send_email_to_user.html', {'user': user})

from django.http import HttpResponse
import csv
from openpyxl import Workbook
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect

@login_required
def export_students(request):
    """Export student data based on filters"""
    if request.user.role != 'instructor':
        return redirect('user_login')
    
    # Get filter parameters
    export_format = request.GET.get('format', 'csv')
    course_id = request.GET.get('course')
    status = request.GET.get('status')
    fields = request.GET.getlist('fields')
    
    # FIXED: Remove 'batch' from select_related - it doesn't exist on Enrollment
    enrollments = Enrollment.objects.filter(
        course__instructor=request.user
    ).select_related('student', 'course')
    
    # Apply filters
    if course_id:
        enrollments = enrollments.filter(course_id=course_id)
    if status:
        enrollments = enrollments.filter(status=status)
    
    # Generate export based on format
    if export_format == 'csv':
        return export_to_csv(enrollments, fields)
    elif export_format == 'excel':
        return export_to_excel(enrollments, fields)
    elif export_format == 'pdf':
        return export_to_pdf(enrollments, fields)
    
    return HttpResponse("Invalid format", status=400)


def export_to_csv(enrollments, fields):
    """Export to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students_export.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    header = ['Student ID', 'Full Name', 'Email']
    if 'contact' in fields:
        header.extend(['Phone Number'])
    if 'enrollment' in fields:
        header.extend(['Course', 'Batch', 'Enrolled Date', 'Status'])
    if 'progress' in fields:
        header.extend(['Progress %'])
    
    writer.writerow(header)
    
    # Write data
    for enrollment in enrollments:
        row = [
            enrollment.student.id,
            enrollment.student.get_full_name(),
            enrollment.student.email
        ]
        
        if 'contact' in fields:
            row.append(enrollment.student.phone_number or 'N/A')
        
        if 'enrollment' in fields:
            # FIXED: Get batch through BatchEnrollment
            try:
                batch_enrollment = enrollment.student.batch_enrollments.filter(
                    batch__course=enrollment.course,
                    is_active=True
                ).first()
                batch_name = batch_enrollment.batch.name if batch_enrollment else 'Direct Enrollment'
            except:
                batch_name = 'Direct Enrollment'
            
            row.extend([
                f"{enrollment.course.course_code}",
                batch_name,
                enrollment.enrolled_at.strftime('%Y-%m-%d'),
                enrollment.get_status_display()
            ])
        
        if 'progress' in fields:
            row.append(f"{enrollment.progress_percentage}%")
        
        writer.writerow(row)
    
    return response


def export_to_excel(enrollments, fields):
    """Export to Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Students Data"
    
    # Write header
    header = ['Student ID', 'Full Name', 'Email']
    if 'contact' in fields:
        header.extend(['Phone Number'])
    if 'enrollment' in fields:
        header.extend(['Course', 'Batch', 'Enrolled Date', 'Status'])
    if 'progress' in fields:
        header.extend(['Progress %'])
    
    ws.append(header)
    
    # Style header
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    
    # Write data
    for enrollment in enrollments:
        row = [
            enrollment.student.id,
            enrollment.student.get_full_name(),
            enrollment.student.email
        ]
        
        if 'contact' in fields:
            row.append(enrollment.student.phone_number or 'N/A')
        
        if 'enrollment' in fields:
            # FIXED: Get batch through BatchEnrollment
            try:
                batch_enrollment = enrollment.student.batch_enrollments.filter(
                    batch__course=enrollment.course,
                    is_active=True
                ).first()
                batch_name = batch_enrollment.batch.name if batch_enrollment else 'Direct Enrollment'
            except:
                batch_name = 'Direct Enrollment'
            
            row.extend([
                f"{enrollment.course.course_code}",
                batch_name,
                enrollment.enrolled_at.strftime('%Y-%m-%d'),
                enrollment.get_status_display()
            ])
        
        if 'progress' in fields:
            row.append(enrollment.progress_percentage)
        
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="students_export.xlsx"'
    wb.save(response)
    
    return response


def export_to_pdf(enrollments, fields):
    """Export to PDF format"""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from io import BytesIO
    except ImportError:
        return HttpResponse("PDF export requires reportlab library. Install it with: pip install reportlab", status=400)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Student Export Report</b>", styles['Heading1'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Prepare data
    data = [['ID', 'Name', 'Email', 'Course', 'Batch', 'Status', 'Progress']]
    
    for enrollment in enrollments:
        # FIXED: Get batch through BatchEnrollment
        try:
            batch_enrollment = enrollment.student.batch_enrollments.filter(
                batch__course=enrollment.course,
                is_active=True
            ).first()
            batch_name = batch_enrollment.batch.name if batch_enrollment else 'Direct'
        except:
            batch_name = 'Direct'
        
        data.append([
            str(enrollment.student.id),
            enrollment.student.get_full_name()[:25],
            enrollment.student.email[:30],
            enrollment.course.course_code,
            batch_name[:15],
            enrollment.get_status_display(),
            f"{enrollment.progress_percentage}%"
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')])
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="students_export.pdf"'
    
    return response